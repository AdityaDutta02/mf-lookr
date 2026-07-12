#!/usr/bin/env python3
"""Deterministic parser for PPFAS's "Detailed Portfolio Disclosure" XLS —
the SEBI-mandated monthly portfolio statement, NOT the marketing factsheet
PDF (see tools/parse_ppfas.py, superseded for months this file covers).

Despite the .xls extension the file is real OOXML (openpyxl needs a .xlsx
copy to load it — see tools/README.md). One sheet per scheme, consistent
columns: Name of the Instrument | ISIN | Industry/Rating | Quantity |
Market/Fair Value (Rs. Lakhs) | % to Net Assets | YTM | YTC.

Row classification is simple and robust: a row with a non-empty ISIN is a
real holding (numbers straight from the cell — no LLM, no OCR). A row with
no ISIN is either a section-label ("Equity & Equity related", "Debt
Instruments", "(a) Listed / awaiting listing...") or a Sub Total/Total/
GRAND TOTAL marker — both skipped as holdings, the label ones update
classification context for rows beneath them. GRAND TOTAL's own weight
(always ~1.0) is a free per-sheet validation check.

Coverage (full-history backfill, 2026-07): 110 of 148 possible months from
2014-03 through 2026-06 are downloadable at all — the remaining 38
(2014-12..2017-11, and 2021-09..2021-10) have NO "Detailed Portfolio
Disclosure" link on PPFAS's own factsheet index page (verified directly
against the index HTML, not a discover-script regex bug: no portfolio-
disclosure URL for those months exists to find). PPFAS simply never
published/archived a Detailed Disclosure XLS for that stretch; the only
thing available for it is the marketing factsheet PDF, which this pipeline
intentionally does not use (see module docstring above and git commit
1ed6d55). Of the 110 available months, all 110 parse, and 413 of 414
fund-months pass the trust-band self-check (holdings sum within [99,101]%
of GRAND TOTAL). The one exception: Parag Parikh Arbitrage Fund, 2024-12,
where the sheet's own "Net Receivables / (Payables)" line (67.6% of NAV)
plus the physical-securities total already exceeds the sheet's own GRAND
TOTAL (100%) — a genuine inconsistency in PPFAS's own source file for that
single month (confirmed: every other Arbitrage Fund month is clean), not a
parser bug. That fund-month is still included in the output (121 real
holdings) with its mismatch left visible rather than silently dropped.

Era-specific quirks this parser reconciles (see inline comments at each
fix for detail):
  - "% to <X>" column scale is NOT determined by its header wording — the
    same header text has meant either a fraction (0.0269 == 2.69%) or an
    already-in-percent value (8.5 == 8.5%) in different eras. Detected
    per-sheet from the data itself (see the "scale" logic in parse_sheet).
  - "GRAND TOTAL" is matched as a prefix, not exact-equality, because some
    eras label it "GRAND TOTAL (AUM)" or "Grand Total (A+B+C+D)".
  - 2017-2019 sheets intersperse futures/currency hedge rows directly into
    the equity listing (not a separate post-GRAND-TOTAL "Derivatives"
    table like later eras) — see INLINE_HEDGE_RE.
  - The 2018 "Sr.No | ISIN | Name | ..." column layout shifts a no-ISIN
    holding's label (CBLO, Net Current Assets) and section/Total labels one
    column left of where data rows keep their name — see the ISIN-column
    fallback and stray-label scan in parse_sheet.
  - Scheme names changed over the years (PPFAS Long Term Value Fund ->
    Parag Parikh Long Term Value Fund -> ... Long Term Equity Fund -> ...
    Flexi Cap Fund; Tax Saver Fund -> ELSS Tax Saver Fund) — canonicalized
    in tools/build_dataset.py's FUND_NAME_ALIASES, not here, since this
    module's job is just to report whatever name the sheet itself used.
"""
import json
import re
import sys
from pathlib import Path

import openpyxl
import xlrd

ROOT = Path(__file__).parent
CACHE_DIR = ROOT / "cache" / "ppfas" / "xlsx"
OUT_DIR = ROOT / "out" / "ppfas_xlsx"

SKIP_RE = re.compile(r"^(sub\s*total|total|grand\s*total)\b", re.IGNORECASE)

SECTION_TYPE_MAP = [
    (re.compile(r"certificate of deposit", re.I), "cd"),
    (re.compile(r"commercial paper", re.I), "cp"),
    (re.compile(r"treasury\s*bill|t-?\s*bills?\b", re.I), "tbill"),
    # "Goverment" (sic) is a real recurring typo in PPFAS's own sheet — gov(?:ern|er)ment
    # matches both "government" and the typo'd "goverment".
    (re.compile(r"gov(?:ern|er)ment (bond|security|securities)|g-?sec|state (govern|gover)ment|state development loan", re.I), "gsec"),
    (re.compile(r"treps|reverse repo|repo", re.I), "treps"),
    (re.compile(r"mutual fund", re.I), "fund"),
    (
        re.compile(
            r"corporate (debt|bond)|debenture|\bncd\b|non.convertible|money market|"
            r"\bdebt instrument|\bdebt securit",
            re.I,
        ),
        "corporate_debt",
    ),
    (re.compile(r"cash|net (current|receivable)", re.I), "cash"),
    (re.compile(r"reit|invit", re.I), "reit"),
    (re.compile(r"future|option|derivative|hedg", re.I), "derivative"),
    (re.compile(r"equity", re.I), "equity"),
]

# 2017-2019 era only: hedge/futures rows (e.g. "BANK OF BARODA-22FEB2018 FUT  #",
# "CUR_USDINR-26MAR2018 FUT  #") are interspersed directly inside the equity
# listing instead of a separate post-GRAND-TOTAL "Derivatives" table (that's how
# later eras do it, and those are already excluded by the GRAND TOTAL break
# below). Their "% to NAV" is hedge notional, not portfolio weight, and PPFAS's
# own "Total"/"Grand Total" rows do NOT include them — confirmed by summing the
# physical-security rows alone against the sheet's own Total row. Detect and
# drop these rows entirely (not just reclassify) or the holdings sum will never
# match GRAND TOTAL.
INLINE_HEDGE_RE = re.compile(r"-\d{1,2}[A-Z]{3}\d{4}\s*FUT\b", re.I)

# Real ISIN: 2-letter country code + 10 alphanumeric chars, e.g. "INE040A01026".
ISIN_RE = re.compile(r"^[A-Z]{2}[A-Z0-9]{10}$")


def classify(section_label: str, name: str = "") -> str:
    # section_label first — this format's section headers are clean and reliable
    # ("Equity & Equity related", "Certificate of Deposit", "Government Securities",
    # ...), so trust it over the row's own name (a company literally named e.g.
    # "Future Retail Limited" would otherwise false-positive-match the "future"
    # derivative pattern). Only fall back to the row's own name for holdings that
    # stand alone with no preceding label row of their own, like "Net Receivables /
    # (Payables)".
    for text in (section_label, name):
        for pat, kind in SECTION_TYPE_MAP:
            if pat.search(text or ""):
                return kind
    return "equity"  # default: most unlabeled-section rows in this format are equity


def num(v):
    if v is None or v == "" or v == "NIL":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


COLUMN_NAME_PATTERNS = {
    # "Name of the Instrument" (most eras) vs "Name of Instrument" (2014 era, no "the").
    "name": re.compile(r"name of (the )?instrument", re.I),
    "isin": re.compile(r"^isin", re.I),
    "industry": re.compile(r"industry|rating", re.I),
    "quantity": re.compile(r"quantity", re.I),
    "market_value": re.compile(r"market.*value", re.I),
    # Header wording drifted across eras: "% to Net Assets" (2023+), "% to AUM"
    # (~2019-2022), "% to NAV" (2014, and again 2019-2022 on some fund sheets).
    # NOTE: the wording alone does NOT tell you the value's *scale* — see the
    # dynamic per-sheet scale detection in parse_sheet() below.
    "weight": re.compile(r"%\s*to\s*(net|aum|nav)", re.I),
}


def load_workbook_rows(path: Path):
    """Returns {sheet_name: [[cell, ...], ...]} regardless of whether the file
    is real OOXML (renamed .xlsx, openpyxl) or legacy BIFF8 (genuine .xls,
    xlrd) — PPFAS serves both under a ".xls" URL depending on how old the
    month is. Detect by the actual file signature, not the extension."""
    with open(path, "rb") as f:
        sig = f.read(4)
    if sig == b"PK\x03\x04":
        wb = openpyxl.load_workbook(str(path), data_only=True)
        return {name: [list(row) for row in wb[name].iter_rows(values_only=True)] for name in wb.sheetnames}
    wb = xlrd.open_workbook(str(path))
    return {name: [wb.sheet_by_name(name).row_values(i) for i in range(wb.sheet_by_name(name).nrows)]
            for name in wb.sheet_names()}


SCHEME_LABEL_RE = re.compile(r"(?:name of the scheme|scheme name)\s*[:\-]\s*(.+)", re.I)


def find_fund_name(rows):
    """Fund name location varies by era (new format: row0 col1; "SCHEME NAME :"
    labeled cell a few rows down in one era; "Name of the Scheme:" labeled cell
    further down still (row ~9) in the 2017-2019 era — that label sits outside
    an 8-row window, and even within a wider window a longer *unrelated* title
    line (e.g. "Monthly Portfolio Statement of the Scheme/s of PPFAS MUTUAL
    FUND as on ...") can out-"longest" it, so an explicit label match always
    wins over the length heuristic when present.
    """
    for r in rows[:15]:
        for c in r:
            if not isinstance(c, str):
                continue
            m = SCHEME_LABEL_RE.search(c)
            if m:
                return re.sub(r"\s*\(.*", "", m.group(1), flags=re.DOTALL).strip()

    # Fallback: no explicit label found — longest cell text containing "Fund"
    # in the first few rows is reliably the scheme's full descriptive name.
    best = None
    for r in rows[:8]:
        for c in r:
            if isinstance(c, str) and "fund" in c.lower() and len(c) > 10:
                if best is None or len(c) > len(best):
                    best = c
    if not best:
        return None
    return re.sub(r"\s*\(.*", "", best, flags=re.DOTALL).strip()


def locate_columns(header_row):
    cols = {}
    for i, cell in enumerate(header_row):
        if not isinstance(cell, str):
            continue
        for key, pat in COLUMN_NAME_PATTERNS.items():
            if key not in cols and pat.search(cell):
                cols[key] = i
    return cols


def parse_sheet(rows):
    """rows: list[list[cell]] — from either openpyxl or xlrd, via load_workbook_rows()."""
    header_idx = next((i for i, r in enumerate(rows) if r and any(
        isinstance(c, str) and COLUMN_NAME_PATTERNS["name"].search(c) for c in r
    )), None)
    if header_idx is None:
        return {"holdings": [], "grand_total": None, "fund_name": None}

    cols = locate_columns(rows[header_idx])
    if "name" not in cols or "weight" not in cols:
        return {"holdings": [], "grand_total": None, "fund_name": None}

    fund_name = find_fund_name(rows)
    if not fund_name:
        return {"holdings": [], "grand_total": None, "fund_name": None}

    holdings = []
    section_label = None
    grand_total_raw = None

    def cell(r, key):
        idx = cols.get(key)
        return r[idx] if idx is not None and idx < len(r) else None

    for r in rows[header_idx + 1:]:
        if not r:
            continue
        raw_name = cell(r, "name")
        name = raw_name.strip() if isinstance(raw_name, str) else raw_name
        shifted_isin = False
        if not name:
            # The "name" column is empty on this row. Two different reasons, both
            # from the same 2018 "Sr.No | ISIN | Name | ..." era layout:
            #
            # 1. A no-ISIN holding (CBLO, Net Current Assets — nothing to put in
            #    the ISIN column) has its label shifted one column LEFT into the
            #    ISIN slot instead of leaving ISIN blank and name in place. If
            #    that cell holds text that isn't itself a real ISIN, it's the
            #    row's actual name.
            raw_isin_cell = cell(r, "isin")
            if isinstance(raw_isin_cell, str) and raw_isin_cell.strip() and not ISIN_RE.match(raw_isin_cell.strip()):
                name = raw_isin_cell.strip()
                shifted_isin = True
            else:
                # 2. Same shift applies to the section/Total/Grand-Total label rows
                #    — scan the whole row for a SKIP_RE label before giving up.
                #    Without this the terminal GRAND TOTAL row is silently missed
                #    and parsing runs on into the Notes/dividend-history section.
                stray_label = next(
                    (c.strip() for c in r if isinstance(c, str) and SKIP_RE.match(c.strip())), None
                )
                if stray_label and stray_label.upper().startswith("GRAND TOTAL"):
                    grand_total_raw = num(cell(r, "weight"))
                    break
                continue
        if isinstance(name, str) and SKIP_RE.match(name):
            # Match "GRAND TOTAL" as a prefix, not exact equality — some eras label
            # it "GRAND TOTAL (AUM)" and an exact-equality check would silently miss
            # the row entirely, letting parsing run on into the Derivatives/hedging
            # disclosure tables below (different column layout -> garbage "holdings").
            if name.strip().upper().startswith("GRAND TOTAL"):
                grand_total_raw = num(cell(r, "weight"))
                # Everything after GRAND TOTAL is a DIFFERENT table (performance/
                # returns/dividend history) whose columns coincidentally hold numeric
                # junk in these same positions — stop here or it gets misread as holdings.
                break
            continue

        if isinstance(name, str) and INLINE_HEDGE_RE.search(name):
            continue

        weight_raw = num(cell(r, "weight"))
        if weight_raw is None:
            # No weight in this row at all -> it's a section-label row (e.g. "Equity &
            # Equity related", "Certificate of Deposit"), not a holding. Some genuine
            # holdings (TREPS, Net Receivables/Payables) have NO isin, so isin presence
            # alone can't be the data-row signal — weight presence is what's reliable.
            #
            # Labels nest two levels deep ("Debt Instruments" -> "(a) Listed / awaiting
            # listing on Stock Exchange") and the sub-label is reused verbatim under BOTH
            # equity and debt top-level sections, so it carries no type information of its
            # own. Only overwrite the tracked label when the new text itself matches a
            # known category — otherwise a meaningless structural sub-header would wipe
            # out the correct top-level label the holdings beneath it actually belong to.
            if isinstance(name, str) and any(pat.search(name) for pat, _ in SECTION_TYPE_MAP):
                section_label = name
            continue

        isin = None if shifted_isin else cell(r, "isin")
        industry = cell(r, "industry")
        industry = industry.strip() if isinstance(industry, str) else (industry or "")
        quantity = num(cell(r, "quantity"))
        market_value_lakhs = num(cell(r, "market_value"))
        holdings.append({
            "name": name.strip() if isinstance(name, str) else str(name),
            "isin": (isin.strip() if isinstance(isin, str) else "") or "",
            "industry": industry,
            "quantity": quantity,
            "market_value_cr": round(market_value_lakhs / 100, 4) if market_value_lakhs is not None else None,
            "_weight_raw": weight_raw,
            "type": classify(section_label, name if isinstance(name, str) else ""),
        })

    # The "% to <X>" column's *scale* isn't determined by its header wording — PPFAS
    # has shipped both a fraction (0.0269 == 2.69%, needs *100) and an already-in-percent
    # (8.5 == 8.5%, used as-is) convention under the exact same header text across
    # different eras (e.g. "% to NAV" means fraction in 2014, but already-percent in
    # 2019-2022 on some fund sheets). Detect it per-sheet from the data itself: a
    # sheet's holdings should sum to roughly 100(%). If the raw values already sum to
    # a plausible percent total (>5), they're already-percent; otherwise they're a
    # fraction and need *100.
    raw_sum = sum(h["_weight_raw"] for h in holdings if h["_weight_raw"] is not None)
    scale = 1 if raw_sum > 5 else 100
    for h in holdings:
        h["weight"] = round(h.pop("_weight_raw") * scale, 4)

    grand_total = None
    if grand_total_raw is not None:
        # Normalize to the same "fraction of 100" contract callers already expect
        # (gt_pct = grand_total * 100), regardless of which scale this sheet used.
        grand_total = grand_total_raw if scale == 100 else grand_total_raw / 100

    return {"holdings": holdings, "grand_total": grand_total, "fund_name": fund_name}


def main():
    period = sys.argv[1] if len(sys.argv) > 1 else None
    xlsx_path = Path(sys.argv[2]) if len(sys.argv) > 2 else None
    if not xlsx_path:
        print("usage: parse_ppfas_xlsx.py <period YYYY-MM> <path to .xlsx>")
        return 1

    sheets = load_workbook_rows(xlsx_path)
    funds = {}
    for sheet_name, rows in sheets.items():
        parsed = parse_sheet(rows)
        if not parsed["fund_name"]:
            continue
        total = round(sum(h["weight"] for h in parsed["holdings"]), 2)
        gt = parsed["grand_total"]
        gt_pct = round(gt * 100, 2) if gt is not None else None
        ok = gt_pct is not None and abs(total - gt_pct) < 0.5
        print(f"  {parsed['fund_name']}: {len(parsed['holdings'])} holdings, "
              f"sum={total:.2f}%, GRAND TOTAL={gt_pct}%, {'OK' if ok else 'MISMATCH'}")
        funds[parsed["fund_name"]] = {"holdings": parsed["holdings"], "grand_total_pct": gt_pct}

    if period:
        OUT_DIR.mkdir(parents=True, exist_ok=True)
        out_path = OUT_DIR / f"{period}.json"
        out_path.write_text(json.dumps(funds, indent=2))
        print(f"Written to {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
