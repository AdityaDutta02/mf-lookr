#!/usr/bin/env python3
"""Deterministic parser for ICICI Prudential Mutual Fund's "Monthly Portfolio
Disclosure" xlsx — the SEBI-mandated monthly portfolio statement, one file
per scheme per month (like HDFC; unlike PPFAS's one-workbook-many-sheets).
Confirmed real OOXML across every scheme type sampled directly (equity —
Value Fund, debt — Corporate Bond Fund and Savings Fund, ETF, FOF).

Layout is a FIXED column mapping, confirmed identical across every sampled
scheme:
  col0 = decorative blank
  col1 = Company/Issuer/Instrument Name — ALSO doubles as the section-label
         text column (same dual-purpose pattern PPFAS/HDFC both have)
  col2 = ISIN
  col3 = Coupon (%) — new column PPFAS/HDFC don't have, unused here
  col4 = Industry/Rating
  col5 = Quantity
  col6 = Exposure/Market Value (Rs. Lakh) — LAKHS; divide by 100 for ₹cr
  col7 = % to Nav — a FRACTION (0.0967 = 9.67%, GRAND-TOTAL-equivalent row
         reads ~0.9999...), same convention as PPFAS, NOT HDFC's
         already-percentage column. *100 needed.
  col8 = Yield of the instrument
  col9 = Yield to Call @ (AT1/Tier2 bonds)

Row classification is the one genuinely new wrinkle vs PPFAS/HDFC: ICICI's
section-header rows are NOT weight-free labels — each one carries its own
section SUBTOTAL in col6/col7 (e.g. "Government Securities" appears both as
a header row with the section's aggregate weight, AND, immediately below, as
the literal instrument name on many individual G-Sec holding rows that DO
have real ISINs — the generic display name is reused per-holding). So
"has a weight" no longer distinguishes a real holding from a section
subtotal the way it did for PPFAS/HDFC.

The actual rule used here: ISIN present -> always a real holding (unambiguous
in every sample checked). ISIN absent -> only counted as a holding if the
name matches a small set of known STANDALONE-leaf instrument names that
never have children of their own (TREPS, Net Current Assets, Cash Margin -
Derivatives, Reverse Repo, and single-line derivative rows marked with
ICICI's own "$$" suffix) — everything else with no ISIN is a pure
section/subtotal label, skipped as a holding but still used to update
classify()'s section-label context (same nested-label protection PPFAS/HDFC
use: only overwrite the tracked label when the new text itself matches a
known category pattern).

Validation anchor: ICICI's version of PPFAS's "GRAND TOTAL" / HDFC's "Grand
Total" is literally "Total Net Assets" (confirmed identical text across all
144 scheme files sampled in one full month) — % to Nav on that row is
~0.9999...  Everything after it is a DIFFERENT table ("Details of Stock
Future / Index Future" annexure, NAV/dividend history, footnotes) with
numeric junk reusing the same column positions — stop scanning there, same
fix PPFAS/HDFC both needed.
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

try:
    import xlrd
except ImportError:  # only needed for the rare legacy-.xls file, if any turn up
    xlrd = None

ROOT = Path(__file__).parent
CACHE_DIR = ROOT / "cache" / "icici" / "xlsx"
OUT_DIR = ROOT / "out" / "icici_xlsx"

# Column positions are located dynamically from the header row, NOT hardcoded
# by index — Fund-of-Funds sheets (Gold ETF FOF, Silver ETF FOF, and every
# "...Active FOF"/"...Omni FOF" scheme) omit the "Coupon" column entirely
# (FOFs don't hold coupon-bearing instruments), which shifts every column
# after it left by one versus equity/debt sheets. Confirmed by direct
# inspection: a fixed COL_WEIGHT=7 silently read the wrong cell (weight
# column) on every FOF file and produced 0 holdings. locate_columns() below
# fixes this the same way PPFAS's parser does.
COL_NAME = 1  # header text itself is always present in this column across every sample
COLUMN_NAME_PATTERNS = {
    "isin": re.compile(r"isin", re.I),
    "coupon": re.compile(r"coupon", re.I),
    "industry": re.compile(r"industry|rating", re.I),
    "quantity": re.compile(r"quantity", re.I),
    "market_value": re.compile(r"market\s*value", re.I),
    "weight": re.compile(r"%\s*to\s*nav", re.I),
}


def locate_columns(header_row):
    cols = {}
    for i, raw in enumerate(header_row):
        if not isinstance(raw, str):
            continue
        cell_text = raw.strip()  # FOF sheets have a stray leading tab on the ISIN header cell
        for key, pat in COLUMN_NAME_PATTERNS.items():
            if key not in cols and pat.search(cell_text):
                cols[key] = i
    return cols

GRAND_TOTAL_RE = re.compile(r"^total net assets\b", re.IGNORECASE)
# No-ISIN rows that ARE real leaf holdings (never have their own children) —
# confirmed by inspecting every no-ISIN, weight-bearing row across a full
# month's 144 files (see tools/README.md-adjacent notes / task report).
#
# "Reverse Repo" is the one exception that needed a second pass: some funds
# (Overnight/Liquid/Savings — the cash-heavy ones) print a bare "Reverse
# Repo" SECTION-SUBTOTAL row (no ISIN) immediately followed by several
# "Reverse Repo (<maturity date>)" CHILD rows that individually sum to that
# same subtotal — sometimes with ISIN-like repo codes (e.g. "GSECREPO1010"),
# sometimes without. Matching bare "Reverse Repo" as a leaf too (as an
# earlier version of this parser did) double-counts the section twice.
# Fix: only the DATED "Reverse Repo (...)" form counts as a leaf; the bare
# "Reverse Repo" row is treated as a pure section label like any other.
#
# "CBLO" (Collateralized Borrowing and Lending Obligation) is the pre-~2018/19
# era's cash-equivalent instrument name — TREPS didn't exist yet as a category
# (RBI discontinued CBLO and introduced TREPS in Oct 2018). Confirmed across
# every 2015-2018 sample checked: "CBLO" always appears as a single bare row
# with its own weight, immediately followed by "Net Current Assets" or
# "Others" (never has child rows of its own, unlike bare "Reverse Repo"
# above) — so it's safe to always treat as a leaf, same as TREPS today.
# Missing this was the single largest source of pre-2019 trust-band
# mismatches (an 18%+ weight gap per fund in some months) since CBLO wasn't
# in this pattern at all and doesn't match any SECTION_TYPE_MAP pattern
# either, so it was silently dropped entirely as neither a leaf nor a
# recognized section label.
STANDALONE_LEAF_RE = re.compile(
    r"^(net current assets|net receivables?|treps|reverse repo\s*\(|cash margin|cblo)", re.IGNORECASE
)
# Individual bank-deposit leaf rows under the "Deposits (Placed as Margin)"
# subsection — pre-2019-era arbitrage/hedged-equity funds only, e.g. "ICICI
# Bank Ltd. - 12 Dec 2016 (Duration - 364 Days)". No ISIN, no Industry/Rating
# value (that column is genuinely blank for these, unlike the "has an
# industry value" tell case 3 relies on for other one-off leaf rows), so
# these were being silently dropped as pure section-subtotal noise. Confirmed
# safe: the parent "Deposits (Placed as Margin)" header row's own text never
# recurs verbatim in these (each child's name embeds its own date), so
# there's no risk of double-counting the header via the existing bare-label
# repeat check above.
DURATION_DEPOSIT_RE = re.compile(r"\(duration\s*-?\s*\d+\s*days?\)", re.IGNORECASE)
DERIVATIVE_MARKER_RE = re.compile(r"\$\$")

SECTION_TYPE_MAP = [
    (re.compile(r"certificate of deposit", re.I), "cd"),
    (re.compile(r"commercial paper", re.I), "cp"),
    (re.compile(r"treasury\s*bills?\b", re.I), "tbill"),
    (re.compile(r"government securit|g-?sec|state government|state development loan", re.I), "gsec"),
    (re.compile(r"treps|reverse repo|cblo", re.I), "treps"),
    (re.compile(r"securitised|securitized", re.I), "securitized"),
    (re.compile(r"alternative investment fund|\baif\b", re.I), "fund"),
    (re.compile(r"mutual fund", re.I), "fund"),
    (re.compile(r"infrastructure investment trust|\binvit", re.I), "invit"),
    (re.compile(r"real estate investment trust|\breit", re.I), "reit"),
    (re.compile(r"foreign securit|overseas etf", re.I), "foreign_equity"),
    (
        re.compile(
            r"non.convertible debentures?|bonds?\b|debentures?|\bncd\b|money market|"
            r"zero coupon|deep discount|debt instrument",
            re.I,
        ),
        "corporate_debt",
    ),
    (re.compile(r"cash margin|net current|net receivable|deposit", re.I), "cash"),
    (re.compile(r"future|option|derivative|hedg", re.I), "derivative"),
    (re.compile(r"preference share", re.I), "preference"),
    (re.compile(r"equity", re.I), "equity"),
]


def classify(section_label: str, name: str = "") -> str:
    for text in (section_label, name):
        for pat, kind in SECTION_TYPE_MAP:
            if pat.search(text or ""):
                return kind
    return "equity"


def num(v):
    if v is None or v == "" or v in ("NIL", "Nil", "nil", "^"):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def load_workbook_sheets(path: Path):
    """Detect by file signature, not extension — same "extension lies"
    precaution PPFAS/HDFC both needed, even though every sample checked here
    was genuine OOXML."""
    with open(path, "rb") as f:
        sig = f.read(8)
    if sig[:4] == b"PK\x03\x04":
        wb = openpyxl.load_workbook(str(path), data_only=True)
        return {name: [list(row) for row in wb[name].iter_rows(values_only=True)] for name in wb.sheetnames}
    if sig[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        if xlrd is None:
            raise RuntimeError(f"{path} is legacy BIFF8/OLE2 but xlrd isn't installed")
        wb = xlrd.open_workbook(str(path))
        return {name: [wb.sheet_by_name(name).row_values(i) for i in range(wb.sheet_by_name(name).nrows)]
                for name in wb.sheet_names()}
    raise RuntimeError(f"{path}: unrecognized file signature {sig!r}")


def find_fund_name(rows):
    """Row 1 (0-indexed), col 1 is the scheme's own name — row 0/col1 is
    always the literal "ICICI Prudential Mutual Fund" AMC banner, confirmed
    across every sample. No parenthetical suffix to strip (unlike HDFC)."""
    if len(rows) < 2 or not rows[1] or len(rows[1]) <= COL_NAME:
        return None
    cell = rows[1][COL_NAME]
    if not isinstance(cell, str) or not cell.strip():
        return None
    return cell.strip()


def cell(r, idx):
    return r[idx] if idx is not None and idx < len(r) else None


def parse_sheet(rows):
    header_idx = next((i for i, r in enumerate(rows) if r and any(
        isinstance(c, str) and re.search(r"company/issuer/instrument name", c, re.I) for c in r
    )), None)
    if header_idx is None:
        return {"holdings": [], "grand_total": None, "fund_name": None}

    cols = locate_columns(rows[header_idx])
    if "isin" not in cols or "weight" not in cols:
        return {"holdings": [], "grand_total": None, "fund_name": None}

    fund_name = find_fund_name(rows)
    if not fund_name:
        return {"holdings": [], "grand_total": None, "fund_name": None}

    holdings = []
    section_label = None
    grand_total = None
    # Tracks the normalized text of the immediately-preceding no-ISIN row —
    # needed for the "Reverse Repo" repeat case below.
    prev_bare_label = None

    for r in rows[header_idx + 1:]:
        if not r:
            continue
        raw_name = cell(r, COL_NAME)
        name = raw_name.strip() if isinstance(raw_name, str) else raw_name
        if not name:
            continue

        if isinstance(name, str) and GRAND_TOTAL_RE.match(name):
            grand_total = num(cell(r, cols["weight"]))
            # Everything after "Total Net Assets" is a different table (stock/
            # index future annexure, NAV history, footnotes) with numeric junk
            # in the same column positions — stop here.
            break

        isin_cell = cell(r, cols["isin"])
        isin = isin_cell.strip() if isinstance(isin_cell, str) else ""
        weight_frac = num(cell(r, cols["weight"]))

        industry_cell = cell(r, cols.get("industry"))
        industry_val = industry_cell.strip() if isinstance(industry_cell, str) else industry_cell

        if not isin:
            normalized = name.lower()
            is_repeat_of_prev_label = normalized == prev_bare_label
            prev_bare_label = normalized
            if is_repeat_of_prev_label:
                # Some months print undated "Reverse Repo" child rows
                # identical in text to their own section-header row (unlike
                # the more common dated "Reverse Repo (<date>)" form the
                # STANDALONE_LEAF_RE regex below already handles) — e.g. a
                # header "Reverse Repo" (weight X) directly followed by
                # several more literal "Reverse Repo" rows whose weights sum
                # to X. There's no textual way to tell header from child
                # here; the structural tell is that a genuine one-off
                # section header never repeats itself verbatim on the very
                # next row. Treat any such immediate repeat as a real leaf.
                if weight_frac is None:
                    continue
                quantity = num(cell(r, cols.get("quantity")))
                market_value_lakhs = num(cell(r, cols.get("market_value")))
                holdings.append({
                    "name": name, "isin": isin, "industry": industry_val or "",
                    "quantity": quantity,
                    "market_value_cr": round(market_value_lakhs / 100, 4) if market_value_lakhs is not None else None,
                    "weight": round(weight_frac * 100, 4),
                    "type": classify(section_label, name),
                })
                continue
            # Three-way call for a no-ISIN row, in priority order:
            #  1. Known standalone leaf (TREPS, Net Current Assets, Cash
            #     Margin, a DATED "Reverse Repo (...)", a "$$"-marked
            #     derivative line) -> always a real holding.
            #  2. Otherwise, if it matches one of the known broad SECTION
            #     patterns (Government Securities, Debt Instruments, ...) ->
            #     it's a section/subtotal header — skip as a holding, use
            #     only for classify() context (e.g. bare "Reverse Repo" lands
            #     here, not in case 1, since STANDALONE_LEAF_RE requires the
            #     dated parenthetical form).
            #  3. Otherwise (matches NEITHER a known leaf NOR a known section
            #     pattern) -> count it as a holding ONLY if it has a
            #     populated Industry/Rating cell — confirmed the reliable
            #     tell for "this is an actual leaf line" vs "this is a
            #     structural subtotal row" across every sampled file: real
            #     holdings (including unclassifiable one-offs like physical
            #     "Gold (995 Purity) ^" / "Silver ^", industry="Gold"/
            #     "Silver") always have SOMETHING there; pure structural
            #     headers ("Listed / Awaiting Listing On Stock Exchanges",
            #     "Others", "Privately Placed/Unlisted") never do, even
            #     though they carry a nonzero subtotal weight of their own.
            is_standalone_leaf = bool(
                STANDALONE_LEAF_RE.match(name) or DERIVATIVE_MARKER_RE.search(name)
                or DURATION_DEPOSIT_RE.search(name)
            )
            if not is_standalone_leaf:
                matches_section_pattern = any(pat.search(name) for pat, _ in SECTION_TYPE_MAP)
                if matches_section_pattern or not industry_val:
                    # Pure section/subtotal label row — update classify()
                    # context (only when the label itself matches a known
                    # pattern — a nested structural sub-header carries no
                    # type info of its own, same guard PPFAS/HDFC use) and
                    # skip as a holding.
                    if matches_section_pattern:
                        section_label = name
                    continue
                # Case 3 fallback (has an industry value) — count as a holding.
            if weight_frac is None:
                continue  # e.g. "Units of an Alternative Investment Fund (AIF)" = "Nil"
        else:
            prev_bare_label = None  # a real ISIN-bearing row breaks any repeat run
            if weight_frac is None:
                continue

        quantity = num(cell(r, cols.get("quantity")))
        market_value_lakhs = num(cell(r, cols.get("market_value")))
        holdings.append({
            "name": name,
            "isin": isin,
            "industry": industry_val or "",
            "quantity": quantity,
            # Lakhs -> Crores.
            "market_value_cr": round(market_value_lakhs / 100, 4) if market_value_lakhs is not None else None,
            # Fraction -> percentage (ICICI's convention matches PPFAS, not HDFC).
            "weight": round(weight_frac * 100, 4),
            "type": classify(section_label, name),
        })

    return {"holdings": holdings, "grand_total": grand_total, "fund_name": fund_name}


def main():
    period = sys.argv[1] if len(sys.argv) > 1 else None
    xlsx_path = Path(sys.argv[2]) if len(sys.argv) > 2 else None
    if not xlsx_path:
        print("usage: parse_icici_xlsx.py <period YYYY-MM> <path to .xlsx>")
        return 1

    sheets = load_workbook_sheets(xlsx_path)
    funds = {}
    for sheet_name, rows in sheets.items():
        if "deriv" in sheet_name.lower():
            continue  # SEBI derivative-disclosure notice sheet, not a holdings table
        parsed = parse_sheet(rows)
        if not parsed["fund_name"]:
            continue
        total = round(sum(h["weight"] for h in parsed["holdings"]), 2)
        gt = parsed["grand_total"]
        gt_pct = round(gt * 100, 2) if gt is not None else None
        ok = gt_pct is not None and abs(total - gt_pct) < 0.5
        print(f"  {parsed['fund_name']}: {len(parsed['holdings'])} holdings, "
              f"sum={total:.2f}%, Total Net Assets={gt_pct}%, {'OK' if ok else 'MISMATCH'}")
        funds[parsed["fund_name"]] = {"holdings": parsed["holdings"], "grand_total_pct": gt_pct}

    if period:
        OUT_DIR.mkdir(parents=True, exist_ok=True)
        out_path = OUT_DIR / f"{period}.json"
        out_path.write_text(json.dumps(funds, indent=2))
        print(f"Written to {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
