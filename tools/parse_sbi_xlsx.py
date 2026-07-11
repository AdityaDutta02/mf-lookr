#!/usr/bin/env python3
"""Deterministic parser for SBI Mutual Fund's "Detailed Portfolio
Disclosure" xlsx — the SEBI-mandated monthly portfolio statement, one file
per scheme (like HDFC's, unlike PPFAS's one-workbook-many-sheets; see
parse_hdfc_xlsx.py). Confirmed real OOXML regardless of URL extension
(.xls or .xlsx) — same "extension lies" pattern as PPFAS/HDFC, detect by
file signature not extension.

Layout, confirmed by inspecting a recent equity/hybrid fund (Balanced
Advantage) AND a recent pure debt fund (Corporate Bond) AND a 2019-era file
(Arbitrage Opportunities):
  row1 col2 = "SBI Mutual Fund" / "SBI MUTUAL FUND"
  row2 col2 = "SCHEME NAME :", col3 = the scheme's actual name (source of
              truth for identity — NEVER the discovery manifest's link-text
              title, which is inconsistent across eras; see
              discover_sbi_xlsx.py's docstring)
  row3 col2 = "PORTFOLIO STATEMENT AS ON :", col3 = date
  header row = "Name of the Instrument / Issuer" | "ISIN" | rating/industry
    (one combined column in recent files: "Rating / Industry^"; TWO split
    columns in 2019-era files: "Rating" then "Industry ^" — confirmed by
    direct inspection of a 2019 file, hence "rating" and "industry" are
    tracked as SEPARATE dynamic columns below and merged at read time,
    whichever is non-empty) | "Quantity" | "Market value (Rs. in Lakhs)" |
    "% to AUM" | "YTM %" | "YTC % ##" | "Notes & Symbols"

Column POSITIONS shift between eras (the Rating/Industry split above moves
every column after it right by one) — so, like PPFAS (and unlike HDFC's
fixed-position format), columns are located dynamically from the header
row's own text, not hardcoded indices.

Row classification: identical "weight presence, not ISIN presence" rule as
PPFAS/HDFC — a row with a value in "% to AUM" is a real holding (TREPS, Net
Receivable/Payable, Margin amount for Derivative positions have no ISIN but
do have a weight); a row with no weight is a section-label row. Same nested-
label fix: only overwrite the tracked section label when the label text
itself matches a known category pattern.

Validation: per-subsection "Total" rows, then "GRAND TOTAL (AUM)" with
% to AUM ~= 100 — same free per-sheet anchor as PPFAS/HDFC's GRAND TOTAL,
except SBI's label has a trailing "(AUM)" so the exact-equality check PPFAS
uses ("GRAND TOTAL") is loosened to a prefix match. Immediately after GRAND
TOTAL comes a separate "DERIVATIVES" table (hedging/non-hedging futures
positions, in the same column positions but a DIFFERENT schema — confirmed
by inspection, e.g. Balanced Advantage Fund May-2026 has 58 derivative rows
after GRAND TOTAL with a "Derivatives Total" of -15.99%) — stop scanning at
GRAND TOTAL, same fix as PPFAS/HDFC needed for their own trailing tables.
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

try:
    import xlrd
except ImportError:  # only needed for a genuine legacy-.xls file, if one turns up
    xlrd = None

ROOT = Path(__file__).parent
CACHE_DIR = ROOT / "cache" / "sbi" / "xlsx"
OUT_DIR = ROOT / "out" / "sbi_xlsx"

SKIP_RE = re.compile(r"^(sub\s*total|total|grand\s*total)\b", re.IGNORECASE)
GRAND_TOTAL_RE = re.compile(r"^grand\s*total\b", re.IGNORECASE)

SECTION_TYPE_MAP = [
    (re.compile(r"certificate of deposit", re.I), "cd"),
    (re.compile(r"commercial paper", re.I), "cp"),
    (re.compile(r"treasury\s*bill|t-?\s*bills?\b", re.I), "tbill"),
    (re.compile(r"gov(?:ern|er)ment (bond|security|securities)|g-?sec|state (govern|gover)ment|state development loan|central government", re.I), "gsec"),
    (re.compile(r"treps|reverse repo|repo", re.I), "treps"),
    (re.compile(r"mutual fund", re.I), "fund"),
    (
        re.compile(
            r"corporate (debt|bond)|debenture|\bncd\b|non.convertible|money market|"
            r"\bdebt instrument|\bdebt securit",
            re.I,
        ),
        "corporate_debt",
    ),
    (re.compile(r"cash|net (current|receivable)|margin amount", re.I), "cash"),
    (re.compile(r"reit|invit", re.I), "reit"),
    (re.compile(r"future|option|derivative|hedg", re.I), "derivative"),
    (re.compile(r"preference share", re.I), "preference"),
    (re.compile(r"equity", re.I), "equity"),
]


def classify(section_label: str, name: str = "") -> str:
    for text in (section_label, name):
        for pat, kind in SECTION_TYPE_MAP:
            if pat.search(text or ""):
                return kind
    return "equity"


def num(v):
    if v is None or v == "" or v == "NIL" or v == "#":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


COLUMN_NAME_PATTERNS = {
    "name": re.compile(r"name of the instrument", re.I),
    "isin": re.compile(r"^isin", re.I),
    # "rating" and "industry" are tracked separately (2019-era files split them into
    # two columns; recent files combine them into one "Rating / Industry^" column,
    # in which case both keys resolve to the same index — see module docstring).
    "rating": re.compile(r"^rating", re.I),
    "industry": re.compile(r"industry", re.I),
    "quantity": re.compile(r"quantity", re.I),
    "market_value": re.compile(r"market.*value", re.I),
    "weight": re.compile(r"%\s*to\s*aum", re.I),
}


def load_workbook_rows(path: Path):
    """Real OOXML regardless of the source URL's extension (.xls or .xlsx) —
    same "extension lies" fix as PPFAS/HDFC; detect by file signature."""
    with open(path, "rb") as f:
        sig = f.read(8)
    if sig[:4] == b"PK\x03\x04":
        wb = openpyxl.load_workbook(str(path), data_only=True)
        return {name: [list(row) for row in wb[name].iter_rows(values_only=True)] for name in wb.sheetnames}
    if sig[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        if xlrd is None:
            raise RuntimeError(f"{path} is legacy BIFF8/OLE2 but xlrd isn't installed")
        wb = xlrd.open_workbook(str(path))
        return {name: [wb.sheet_by_name(name).row_values(i) for i in range(wb.sheet_by_name(name).nrows)]
                for name in wb.sheet_names()}
    raise RuntimeError(f"{path}: unrecognized file signature {sig!r}")


def find_fund_name(rows):
    """"SCHEME NAME :" is a fixed, reliable label cell in every era sampled —
    take the next non-empty cell in that same row as the canonical name."""
    for r in rows[:6]:
        for i, c in enumerate(r):
            if isinstance(c, str) and "scheme name" in c.lower():
                for later in r[i + 1:]:
                    if isinstance(later, str) and later.strip():
                        return later.strip()
    return None


def locate_columns(header_row):
    cols = {}
    for i, cell in enumerate(header_row):
        if not isinstance(cell, str):
            continue
        for key, pat in COLUMN_NAME_PATTERNS.items():
            if key not in cols and pat.search(cell):
                cols[key] = i
    return cols


def parse_sheet(rows):
    header_idx = next((i for i, r in enumerate(rows) if r and any(
        isinstance(c, str) and COLUMN_NAME_PATTERNS["name"].search(c) for c in r
    )), None)
    if header_idx is None:
        return {"holdings": [], "grand_total": None, "fund_name": None}

    cols = locate_columns(rows[header_idx])
    if "name" not in cols or "weight" not in cols:
        return {"holdings": [], "grand_total": None, "fund_name": None}

    fund_name = find_fund_name(rows)
    if not fund_name:
        return {"holdings": [], "grand_total": None, "fund_name": None}

    holdings = []
    section_label = None
    grand_total = None

    def cell(r, key):
        idx = cols.get(key)
        return r[idx] if idx is not None and idx < len(r) else None

    for r in rows[header_idx + 1:]:
        if not r:
            continue
        raw_name = cell(r, "name")
        name = raw_name.strip() if isinstance(raw_name, str) else raw_name
        if not name:
            continue
        if isinstance(name, str) and SKIP_RE.match(name):
            if GRAND_TOTAL_RE.match(name.strip()):
                grand_total = num(cell(r, "weight"))
                # Everything after GRAND TOTAL (AUM) is a separate DERIVATIVES table
                # (hedging/non-hedging futures) in the same column positions but a
                # different schema — stop here or it gets misread as holdings, same
                # fix PPFAS/HDFC needed for their own trailing tables.
                break
            continue

        weight_frac = num(cell(r, "weight"))
        if weight_frac is None:
            # No weight -> section-label row, not a holding (see PPFAS's identical
            # reasoning). Only overwrite the tracked label when the new text itself
            # matches a known category, so a meaningless nested sub-header ("a)
            # Listed/awaiting listing...") never wipes out the real label above it.
            if isinstance(name, str) and any(pat.search(name) for pat, _ in SECTION_TYPE_MAP):
                section_label = name
            continue

        isin = cell(r, "isin")
        # Rating and Industry may be the SAME column (recent combined header) or two
        # DIFFERENT columns (2019-era split header) — prefer industry text, fall back
        # to rating, so both eras end up with one sensible "sector" string.
        industry_cell = cell(r, "industry")
        rating_cell = cell(r, "rating")
        industry_cell = industry_cell.strip() if isinstance(industry_cell, str) else (industry_cell or "")
        rating_cell = rating_cell.strip() if isinstance(rating_cell, str) else (rating_cell or "")
        industry = industry_cell or rating_cell
        quantity = num(cell(r, "quantity"))
        market_value_lakhs = num(cell(r, "market_value"))
        holdings.append({
            "name": name.strip() if isinstance(name, str) else str(name),
            "isin": (isin.strip() if isinstance(isin, str) else "") or "",
            "industry": industry,
            "quantity": quantity,
            # Lakhs -> Crores, same conversion PPFAS/HDFC need for their own
            # Lakhs-denominated market-value column.
            "market_value_cr": round(market_value_lakhs / 100, 4) if market_value_lakhs is not None else None,
            # "% to AUM" is already a percentage (GRAND TOTAL (AUM) = 100, not 1.0) —
            # same as HDFC's "% to NAV", unlike PPFAS's source (a fraction).
            "weight": round(weight_frac, 4),
            "type": classify(section_label, name if isinstance(name, str) else ""),
        })

    return {"holdings": holdings, "grand_total": grand_total, "fund_name": fund_name}


def main():
    period = sys.argv[1] if len(sys.argv) > 1 else None
    xlsx_path = Path(sys.argv[2]) if len(sys.argv) > 2 else None
    if not xlsx_path:
        print("usage: parse_sbi_xlsx.py <period YYYY-MM> <path to .xlsx>")
        return 1

    sheets = load_workbook_rows(xlsx_path)
    funds = {}
    for sheet_name, rows in sheets.items():
        parsed = parse_sheet(rows)
        if not parsed["fund_name"]:
            continue
        total = round(sum(h["weight"] for h in parsed["holdings"]), 2)
        gt = parsed["grand_total"]
        gt_pct = round(gt, 2) if gt is not None else None
        ok = gt_pct is not None and abs(total - gt_pct) < 0.5
        print(f"  {parsed['fund_name']}: {len(parsed['holdings'])} holdings, "
              f"sum={total:.2f}%, GRAND TOTAL={gt_pct}%, {'OK' if ok else 'MISMATCH'}")
        funds[parsed["fund_name"]] = {"holdings": parsed["holdings"], "grand_total_pct": gt_pct}

    if period:
        OUT_DIR.mkdir(parents=True, exist_ok=True)
        out_path = OUT_DIR / f"{period}.json"
        out_path.write_text(json.dumps(funds, indent=2))
        print(f"Written to {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
