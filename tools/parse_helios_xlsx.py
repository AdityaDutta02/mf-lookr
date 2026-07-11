#!/usr/bin/env python3
"""Deterministic parser for Helios Mutual Fund's "Monthly Portfolio" XLSX —
the SEBI-mandated "Detailed Portfolio Disclosure" statement, NOT the
marketing factsheet PDF under the site's "Factsheets" tab (same PDF-vs-XLS
distinction as PPFAS/HDFC — see tools/parse_ppfas_xlsx.py's docstring).

One file per (scheme, month) — unlike PPFAS's single workbook with one sheet
per scheme, each Helios download is its own workbook with exactly one sheet
(named a short scheme code, e.g. "HSCF", "HOF", "HARF" — not used, we read
by position/content instead). Real OOXML throughout (openpyxl loads every
sample fine, no xlrd/.xls-disguised-as-.xlsx fallback needed like PPFAS's
older years).

Column layout closely mirrors PPFAS's format (same column-name-based
`locate_columns`/COLUMN_NAME_PATTERNS approach reused verbatim) with THREE
Helios-specific quirks, confirmed against real downloaded files before
writing this:

1. Weight is stored as an actual PERCENTAGE already (e.g. 2.24 meaning
   2.24%), not a 0..1 fraction like PPFAS's "% to Net Assets" column — the
   GRAND TOTAL row itself reads ~100, not ~1.0. Do NOT multiply by 100 (that
   would be the PPFAS behavior and is wrong here).
2. The total-anchor row is literally "GRAND TOTAL (AUM)", not a bare "GRAND
   TOTAL" — matched by prefix, not exact equality.
3. There's an extra unlabeled column B (an internal AMC instrument code,
   e.g. "101370") between the blank column A and the "Name of the
   Instrument" column C — harmless since columns are located by header text,
   not fixed position, so it's simply never referenced.

Hybrid schemes (Arbitrage, Balanced Advantage) carry a SECOND holdings-like
table for derivative/hedge positions (a "DERIVATIVES" section with its own
"Name of the Instrument" sub-header, no ISIN) that starts *after* the GRAND
TOTAL row. Like PPFAS, parsing stops the instant "GRAND TOTAL" is seen, so
that trailing table (short futures with negative weights — not real
portfolio composition) is never touched.

Row classification: identical "weight-presence = data row" rule as PPFAS —
a row with a numeric weight is a holding, a row with none is a section-label
row that updates classification context for holdings beneath it, WHOLLY
skipped as a holding itself. Nested sub-headers only overwrite the tracked
label when the sub-header text itself matches a known instrument-type
pattern (e.g. "Treasury Bills" under "MONEY MARKET INSTRUMENTS") — otherwise
the enclosing top-level label survives, exactly as in PPFAS's classify().
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).parent
CACHE_DIR = ROOT / "cache" / "helios" / "xlsx"
OUT_DIR = ROOT / "out" / "helios_xlsx"

SKIP_RE = re.compile(r"^(sub\s*total|total|grand\s*total)\b", re.IGNORECASE)
GRAND_TOTAL_RE = re.compile(r"^grand\s*total\b", re.IGNORECASE)

SECTION_TYPE_MAP = [
    (re.compile(r"certificate of deposit", re.I), "cd"),
    (re.compile(r"commercial paper", re.I), "cp"),
    (re.compile(r"treasury\s*bill|t-?\s*bills?\b", re.I), "tbill"),
    (re.compile(r"gov(?:ern|er)ment (bond|security|securities)|g-?sec|state (govern|gover)ment|state development loan", re.I), "gsec"),
    (re.compile(r"treps|reverse repo|repo", re.I), "treps"),
    (re.compile(r"mutual fund", re.I), "fund"),
    (
        re.compile(
            r"corporate (debt|bond)|debenture|\bncd\b|non.convertible|money market instrument|"
            r"\bdebt instrument|\bdebt securit",
            re.I,
        ),
        "corporate_debt",
    ),
    # Helios-specific: "Other Current Assets / (Liabilities)" is its own section header
    # (PPFAS's equivalent line is a bare holding row, not a header, so PPFAS's cash
    # pattern never needed to match a *section label* this way) — added so the label
    # row itself sets classification context for "Margin amount for Derivative
    # positions" / "Net Receivable / Payable" beneath it.
    (re.compile(r"other current assets|cash|net (current|receivable)", re.I), "cash"),
    (re.compile(r"reit|invit|real estate investment trust", re.I), "reit"),
    (re.compile(r"future|option|derivative|hedg", re.I), "derivative"),
    (re.compile(r"equity", re.I), "equity"),
]


def classify(section_label: str, name: str = "") -> str:
    for text in (section_label, name):
        for pat, kind in SECTION_TYPE_MAP:
            if pat.search(text or ""):
                return kind
    return "equity"


def num(v):
    if v is None or v == "" or v == "NIL":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


COLUMN_NAME_PATTERNS = {
    "name": re.compile(r"name of the instrument", re.I),
    "isin": re.compile(r"^isin", re.I),
    "industry": re.compile(r"industry|rating", re.I),
    "quantity": re.compile(r"quantity", re.I),
    "market_value": re.compile(r"market.*value", re.I),
    "weight": re.compile(r"%\s*to\s*(net|aum)", re.I),
}


def load_workbook_rows(path: Path):
    """Every sampled Helios file is real OOXML regardless of the .xls/.xlsx
    extension on disk (same "check the signature, not the extension" caution
    as PPFAS, kept here in case an older archive month serves genuine BIFF8)."""
    with open(path, "rb") as f:
        sig = f.read(4)
    if sig == b"PK\x03\x04":
        wb = openpyxl.load_workbook(str(path), data_only=True)
        return {name: [list(row) for row in wb[name].iter_rows(values_only=True)] for name in wb.sheetnames}
    import xlrd
    wb = xlrd.open_workbook(str(path))
    return {name: [wb.sheet_by_name(name).row_values(i) for i in range(wb.sheet_by_name(name).nrows)]
            for name in wb.sheet_names()}


def find_fund_name(rows):
    """Fund name sits in a 'SCHEME NAME :' labeled row (row ~3), formatted as
    "Helios <X> Fund (<long descriptive category blurb>)" — same
    longest-cell-containing-'fund' heuristic as PPFAS's find_fund_name,
    reused verbatim since it already handles the parenthetical strip."""
    best = None
    for r in rows[:8]:
        for c in r:
            if isinstance(c, str) and "fund" in c.lower() and len(c) > 10:
                if best is None or len(c) > len(best):
                    best = c
    if not best:
        return None
    # Split on the first "(" rather than a `\(.*$` regex — some periods' descriptive
    # blurb wraps across an embedded newline inside the cell (e.g. "...large cap, \nmid
    # cap..."), and `.` doesn't cross newlines, silently leaving the whole parenthetical
    # (and the newline) stuck on the name for those months. A plain split is immune to
    # that and also normalizes any other embedded whitespace/newlines via strip().
    return best.split("(", 1)[0].strip()


def locate_columns(header_row):
    cols = {}
    for i, cell in enumerate(header_row):
        if not isinstance(cell, str):
            continue
        for key, pat in COLUMN_NAME_PATTERNS.items():
            if key not in cols and pat.search(cell):
                cols[key] = i
    return cols


def parse_sheet(rows):
    header_idx = next((i for i, r in enumerate(rows) if r and any(
        isinstance(c, str) and COLUMN_NAME_PATTERNS["name"].search(c) for c in r
    )), None)
    if header_idx is None:
        return {"holdings": [], "grand_total": None, "fund_name": None}

    cols = locate_columns(rows[header_idx])
    if "name" not in cols or "weight" not in cols:
        return {"holdings": [], "grand_total": None, "fund_name": None}

    fund_name = find_fund_name(rows)
    if not fund_name:
        return {"holdings": [], "grand_total": None, "fund_name": None}

    holdings = []
    section_label = None
    grand_total = None

    def cell(r, key):
        idx = cols.get(key)
        return r[idx] if idx is not None and idx < len(r) else None

    for r in rows[header_idx + 1:]:
        if not r:
            continue
        raw_name = cell(r, "name")
        name = raw_name.strip() if isinstance(raw_name, str) else raw_name
        if not name:
            continue
        if isinstance(name, str) and SKIP_RE.match(name):
            if GRAND_TOTAL_RE.match(name.strip()):
                # Helios' own "% to AUM" column is already a percentage (not a 0..1
                # fraction like PPFAS) — GRAND TOTAL (AUM) reads ~100 directly.
                grand_total = num(cell(r, "weight"))
                # Everything after GRAND TOTAL in hybrid schemes (Arbitrage, Balanced
                # Advantage) is a SEPARATE "DERIVATIVES" table of hedge positions with
                # its own differently-shaped header row — stop here, exactly like PPFAS.
                break
            continue

        weight_pct = num(cell(r, "weight"))
        if weight_pct is None:
            if isinstance(name, str) and any(pat.search(name) for pat, _ in SECTION_TYPE_MAP):
                section_label = name
            continue

        isin = cell(r, "isin")
        industry = cell(r, "industry")
        industry = industry.strip() if isinstance(industry, str) else (industry or "")
        quantity = num(cell(r, "quantity"))
        market_value_lakhs = num(cell(r, "market_value"))
        holdings.append({
            "name": name.strip() if isinstance(name, str) else str(name),
            "isin": (isin.strip() if isinstance(isin, str) else "") or "",
            "industry": industry,
            "quantity": quantity,
            "market_value_cr": round(market_value_lakhs / 100, 4) if market_value_lakhs is not None else None,
            "weight": round(weight_pct, 4),
            "type": classify(section_label, name if isinstance(name, str) else ""),
        })

    return {"holdings": holdings, "grand_total": grand_total, "fund_name": fund_name}


def main():
    period = sys.argv[1] if len(sys.argv) > 1 else None
    xlsx_path = Path(sys.argv[2]) if len(sys.argv) > 2 else None
    if not xlsx_path:
        print("usage: parse_helios_xlsx.py <period YYYY-MM> <path to .xlsx>")
        return 1

    sheets = load_workbook_rows(xlsx_path)
    funds = {}
    for sheet_name, rows in sheets.items():
        parsed = parse_sheet(rows)
        if not parsed["fund_name"]:
            continue
        total = round(sum(h["weight"] for h in parsed["holdings"]), 2)
        gt_pct = round(parsed["grand_total"], 2) if parsed["grand_total"] is not None else None
        ok = gt_pct is not None and abs(total - gt_pct) < 0.5
        print(f"  {parsed['fund_name']}: {len(parsed['holdings'])} holdings, "
              f"sum={total:.2f}%, GRAND TOTAL={gt_pct}%, {'OK' if ok else 'MISMATCH'}")
        funds[parsed["fund_name"]] = {"holdings": parsed["holdings"], "grand_total_pct": gt_pct}

    if period:
        OUT_DIR.mkdir(parents=True, exist_ok=True)
        out_path = OUT_DIR / f"{period}.json"
        existing = json.loads(out_path.read_text()) if out_path.exists() else {}
        existing.update(funds)
        out_path.write_text(json.dumps(existing, indent=2))
        print(f"Written to {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
