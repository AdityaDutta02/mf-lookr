#!/usr/bin/env python3
"""Deterministic parser for Nippon India Mutual Fund's "Monthly Portfolio
Statement" XLS — the SEBI-mandated monthly portfolio disclosure, NOT the
marketing factsheet PDF (same PDF-vs-XLS distinction as PPFAS/HDFC).

One workbook per month, one sheet per scheme (same shape as PPFAS, unlike
HDFC's one-file-per-scheme-per-month) — a leading "Index" sheet maps each
sheet's short code (e.g. "GF") to the scheme's full descriptive name, but
each scheme sheet ALSO repeats its own code + full name at row0 (cols 0/1),
so per-sheet parsing needs no cross-sheet lookup — same self-contained-sheet
property PPFAS's format has.

Columns are a fixed, header-detected layout confirmed identical from the
oldest sampled file (October 2018, genuine legacy BIFF8) through the newest
(June 2026, real OOXML): a blank/internal-code col0, then ISIN | Name of the
Instrument | Industry / Rating | Quantity | Market/Fair Value (Rs. in Lacs) |
% to NAV | YIELD. Detected by header text (like parse_ppfas_xlsx.py), not
fixed indices, in case an older pre-2018 file (outside the sampled range)
shifts something.

Despite the ".xls" extension, files split across BOTH real formats depending
on era — modern OOXML zip (confirmed 2020+) and legacy BIFF8/OLE2 (confirmed
2018) — same "extension lies" situation PPFAS's and (one HDFC month's) file
had. Detect by file signature, not extension (see load_workbook_rows()).

Row classification: identical "weight presence, not ISIN presence" rule as
PPFAS/HDFC — a row with a numeric % to NAV is a real holding (TREPS, Net
Current Assets, Cash Margin rows have no ISIN but do have a weight). A row
with no weight is a section-label row ("Equity & Equity related", "(a)
Listed / awaiting listing...", "Money Market Instruments", ...) — same
nested-label fix as PPFAS/HDFC: only overwrite the tracked section label
when the label text itself matches a known category pattern, since
structural sub-headers ("(a) Listed...", "Subtotal") are reused verbatim
across sections and carry no type information of their own.

Nippon-specific quirk: rows below roughly 0.01% of NAV sometimes carry the
literal string "$0.00%" in the weight cell instead of a number (confirmed:
"Cash Margin - CCIL" rows) — num() strips a leading "$" and trailing "%"
before float() so these tiny holdings aren't dropped as false section labels.

Validation: same GRAND TOTAL-anchored trust band as PPFAS/HDFC — its own
weight is always ~1.0 (=100%), and everything after it is a DIFFERENT table
(derivative disclosure notes, NAV-per-plan tables, riskometer notes) with
numeric junk in the same column positions — stop scanning there or it gets
misread as holdings, same fix as PPFAS/HDFC.
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

try:
    import xlrd
except ImportError:  # only needed for legacy BIFF8/OLE2 files
    xlrd = None

ROOT = Path(__file__).parent
CACHE_DIR = ROOT / "cache" / "nippon" / "xlsx"
OUT_DIR = ROOT / "out" / "nippon_xlsx"

SKIP_RE = re.compile(r"^(sub\s*total|total|grand\s*total)\b", re.IGNORECASE)

SECTION_TYPE_MAP = [
    (re.compile(r"certificate of deposit", re.I), "cd"),
    (re.compile(r"commercial paper", re.I), "cp"),
    (re.compile(r"treasury\s*bill|t-?\s*bills?\b", re.I), "tbill"),
    (re.compile(r"gov(?:ern|er)ment (bond|security|securities)|g-?sec|state (govern|gover)ment|state development loan|\bsdl\b", re.I), "gsec"),
    (re.compile(r"triparty repo|treps|reverse repo|repo", re.I), "treps"),
    # "Exchange Traded Funds" / "Alternative Investment Fund Units" sections hold
    # units of OTHER funds (e.g. a Gold Savings FoF holding Gold BeES units) —
    # same "fund" bucket as "Mutual Fund Units", checked before the standalone
    # "equity" fallback since ETF units aren't the scheme's own equity book.
    (re.compile(r"mutual fund units?|exchange traded funds?|alternative investment fund", re.I), "fund"),
    (
        re.compile(
            r"corporate (debt|bond)|debenture|\bncd\b|non.convertible|money market|"
            r"\bdebt instrument|\bdebt securit|zero coupon|floating rate note",
            re.I,
        ),
        "corporate_debt",
    ),
    (re.compile(r"cash margin|cash|net current|net receivable", re.I), "cash"),
    (re.compile(r"\breit\b|invit", re.I), "reit"),
    (re.compile(r"future|option|derivative|hedg|interest rate swap", re.I), "derivative"),
    (re.compile(r"preference share", re.I), "preference"),
    # Overseas holdings (Japan/US/Taiwan equity themed schemes, or a "Listed
    # Foreign Securities/Overseas ETFs" section within a multi-asset scheme) are
    # equity exposure for asset-allocation purposes, just not domestic-listed.
    (re.compile(r"listed foreign securit|overseas etf", re.I), "equity"),
    (re.compile(r"equity", re.I), "equity"),
]


def classify(section_label: str, name: str = "") -> str:
    for text in (section_label, name):
        for pat, kind in SECTION_TYPE_MAP:
            if pat.search(text or ""):
                return kind
    return "equity"


def num(v):
    if v is None or v == "" or v == "NIL":
        return None
    if isinstance(v, str):
        # Rows below ~0.01% of NAV sometimes hold the literal string "$0.00%"
        # instead of a number in the weight cell (confirmed: "Cash Margin - CCIL"
        # rows) — strip the currency/percent decoration before float().
        s = v.strip().lstrip("$").rstrip("%").strip()
        if not s or s.upper() == "NIL":
            return None
        try:
            return float(s)
        except ValueError:
            return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


COLUMN_NAME_PATTERNS = {
    "name": re.compile(r"name of the instrument", re.I),
    "isin": re.compile(r"^isin", re.I),
    "industry": re.compile(r"industry|rating", re.I),
    "quantity": re.compile(r"quantity", re.I),
    "market_value": re.compile(r"market.*value", re.I),
    "weight": re.compile(r"%\s*to\s*(net|aum|nav)", re.I),
}


def load_workbook_rows(path: Path):
    """Nippon serves BOTH real OOXML (confirmed 2020+) and legacy BIFF8/OLE2
    (confirmed October 2018) under the same ".xls" URL/extension depending on
    era — same "extension lies" fix as parse_ppfas_xlsx.py/parse_hdfc_xlsx.py.
    Detect by file signature, not extension."""
    with open(path, "rb") as f:
        sig = f.read(8)
    if sig[:4] == b"PK\x03\x04":
        wb = openpyxl.load_workbook(str(path), data_only=True)
        return {name: [list(row) for row in wb[name].iter_rows(values_only=True)] for name in wb.sheetnames}
    if sig[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        if xlrd is None:
            raise RuntimeError(f"{path} is legacy BIFF8/OLE2 but xlrd isn't installed")
        wb = xlrd.open_workbook(str(path))
        return {name: [wb.sheet_by_name(name).row_values(i) for i in range(wb.sheet_by_name(name).nrows)]
                for name in wb.sheet_names()}
    raise RuntimeError(f"{path}: unrecognized file signature {sig!r}")


def find_fund_name(rows):
    """row0/col1 is always the scheme's full descriptive name + a trailing
    parenthetical SEBI scheme-type description, confirmed identical across
    every sampled sheet/era (e.g. "Nippon India Growth Mid Cap Fund (Mid Cap
    Fund- An open ended equity scheme...)"). Strip the parenthetical."""
    if not rows or not rows[0] or len(rows[0]) < 2:
        return None
    cell = rows[0][1]
    if not isinstance(cell, str) or not cell.strip():
        return None
    name = re.sub(r"\s*[\(\[].*$", "", cell, flags=re.S).strip()
    # Some scheme titles wrap onto a second line in the source cell (e.g. "Nippon
    # India Nifty Auto \nIndex Fund") — collapse any embedded whitespace/newlines
    # to a single space so this canonical name is stable and joinable elsewhere.
    return re.sub(r"\s+", " ", name).strip()


def locate_columns(header_row):
    cols = {}
    for i, cell in enumerate(header_row):
        if not isinstance(cell, str):
            continue
        for key, pat in COLUMN_NAME_PATTERNS.items():
            if key not in cols and pat.search(cell):
                cols[key] = i
    return cols


def parse_sheet(rows):
    header_idx = next((i for i, r in enumerate(rows) if r and any(
        isinstance(c, str) and COLUMN_NAME_PATTERNS["name"].search(c) for c in r
    )), None)
    if header_idx is None:
        return {"holdings": [], "grand_total": None, "fund_name": None}

    cols = locate_columns(rows[header_idx])
    if "name" not in cols or "weight" not in cols:
        return {"holdings": [], "grand_total": None, "fund_name": None}

    fund_name = find_fund_name(rows)
    if not fund_name:
        return {"holdings": [], "grand_total": None, "fund_name": None}

    holdings = []
    section_label = None
    grand_total = None

    def cell(r, key):
        idx = cols.get(key)
        return r[idx] if idx is not None and idx < len(r) else None

    for r in rows[header_idx + 1:]:
        if not r:
            continue
        raw_name = cell(r, "name")
        name = raw_name.strip() if isinstance(raw_name, str) else raw_name
        raw_isin = cell(r, "isin")
        isin_val = (raw_isin.strip() if isinstance(raw_isin, str) else raw_isin) or ""
        if not name:
            # Nippon's own source data has at least one confirmed row (a
            # Certificate of Deposit, "BARC34", June 2021 Liquid Fund) with a
            # genuinely blank Name-of-the-Instrument cell despite a real ISIN +
            # weight + market value — falling back to the ISIN as the display
            # name keeps this holding in the AUM/weight totals instead of
            # silently dropping it (confirmed root cause of an isolated
            # GRAND-TOTAL-vs-sum mismatch for that one month/fund).
            if isin_val:
                name = isin_val
            else:
                continue
        if isinstance(name, str) and SKIP_RE.match(name):
            if name.strip().upper() == "GRAND TOTAL":
                grand_total = num(cell(r, "weight"))
                # Everything after GRAND TOTAL is a DIFFERENT table (derivative
                # disclosure notes, per-plan NAV tables, riskometer notes) whose
                # columns coincidentally hold numeric/text junk in these same
                # positions — stop here or it gets misread as holdings.
                break
            continue

        weight_frac = num(cell(r, "weight"))
        if weight_frac is None:
            # No parseable weight -> section-label row, not a holding. Only
            # overwrite the tracked label when the new text itself matches a
            # known category — nested structural sub-headers ("(a) Listed /
            # awaiting listing...", "Subtotal") are reused verbatim across
            # sections and carry no type information of their own.
            if isinstance(name, str) and any(pat.search(name) for pat, _ in SECTION_TYPE_MAP):
                section_label = name
            continue

        industry = cell(r, "industry")
        industry = industry.strip() if isinstance(industry, str) else (industry or "")
        quantity = num(cell(r, "quantity"))
        market_value_lacs = num(cell(r, "market_value"))
        holdings.append({
            "name": name.strip() if isinstance(name, str) else str(name),
            "isin": isin_val,
            "industry": industry,
            "quantity": quantity,
            # Lacs -> Crores (Nippon discloses in Lacs, like HDFC; PPFAS's
            # source is already Crores).
            "market_value_cr": round(market_value_lacs / 100, 4) if market_value_lacs is not None else None,
            "weight": round(weight_frac * 100, 4),
            "type": classify(section_label, name if isinstance(name, str) else ""),
        })

    return {"holdings": holdings, "grand_total": grand_total, "fund_name": fund_name}


def main():
    period = sys.argv[1] if len(sys.argv) > 1 else None
    xlsx_path = Path(sys.argv[2]) if len(sys.argv) > 2 else None
    if not xlsx_path:
        print("usage: parse_nippon_xlsx.py <period YYYY-MM> <path to .xlsx>")
        return 1

    sheets = load_workbook_rows(xlsx_path)
    funds = {}
    for sheet_name, rows in sheets.items():
        parsed = parse_sheet(rows)
        if not parsed["fund_name"]:
            continue
        total = round(sum(h["weight"] for h in parsed["holdings"]), 2)
        gt = parsed["grand_total"]
        gt_pct = round(gt * 100, 2) if gt is not None else None
        ok = gt_pct is not None and abs(total - gt_pct) < 0.5
        print(f"  {parsed['fund_name']}: {len(parsed['holdings'])} holdings, "
              f"sum={total:.2f}%, GRAND TOTAL={gt_pct}%, {'OK' if ok else 'MISMATCH'}")
        funds[parsed["fund_name"]] = {"holdings": parsed["holdings"], "grand_total_pct": gt_pct}

    if period:
        OUT_DIR.mkdir(parents=True, exist_ok=True)
        out_path = OUT_DIR / f"{period}.json"
        out_path.write_text(json.dumps(funds, indent=2))
        print(f"Written to {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
