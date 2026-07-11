#!/usr/bin/env python3
"""Deterministic parser for HDFC Mutual Fund's "Detailed Portfolio Disclosure"
xlsx — the SEBI-mandated monthly portfolio statement, one file per scheme
(unlike PPFAS's one-workbook-many-sheets; see parse_ppfas_xlsx.py). Confirmed
real OOXML across every scheme type sampled (equity, ETF, index fund, FMP,
FOF, pure debt) — openpyxl handles all of them directly, no legacy-.xls path
like PPFAS needed.

Layout is a FIXED column mapping, confirmed identical across an equity fund
(Flexi Cap), two index/ETF trackers (Nifty 50 Index Fund, Nifty 50 ETF), a
pure debt fund (Corporate Bond Fund), an FMP, and two FOFs (Gold ETF FOF,
Multi-Asset Active FOF):
  col0 = decorative "|" marker (ignore)
  col1 = ISIN — ALSO doubles as the section-label text column (same
         dual-purpose pattern as PPFAS's format: a row with a label in this
         column and nothing elsewhere is a header, not a holding)
  col2 = Coupon (%)
  col3 = Name Of the Instrument
  col4 = Industry+ /Rating
  col5 = Quantity
  col6 = Market/Fair Value (Rs. in Lacs.) — LACS, not Crores like PPFAS;
         divide by 100 to get the ₹cr unit the app's Holding.market_value
         and AnalyseData.aum fields expect everywhere.
  col7 = % to NAV
  col8 = Yield
  col9 = ~YTC (AT1/Tier 2 bonds)

Row classification: identical "weight presence, not ISIN presence" rule as
PPFAS (col7 non-empty AND col3 non-empty = real holding; TREPS/Net Current
Assets rows have no ISIN but do have a weight). Section headers ("EQUITY &
EQUITY RELATED", "DEBT INSTRUMENTS", "(a) Listed / awaiting listing...",
"TREPS - Tri-party Repo", "MONEY MARKET INSTRUMENTS", "OTHERS", "Mutual Fund
Units", "Net Current Assets") live in col1 with every other column empty.
Same nested-label rule as PPFAS: only overwrite the tracked section label
when the new label text itself matches a known category pattern — a
structural sub-header ("(a) Listed / awaiting listing...") is reused
verbatim across sections and carries no type information of its own.

Validation: "Sub Total" (per sub-section), "Total" (per top-level section),
then "Grand Total" with % to NAV ~= 100 — same free per-sheet anchor as
PPFAS's GRAND TOTAL. Everything after Grand Total is a DIFFERENT table (Top
Ten Holdings / NAV history / riskometer notes) whose columns coincidentally
hold numeric junk in the same positions — confirmed by inspection (e.g. a
stray 'Cash, Cash Equivalents and Net Current Assets : 3.08' row at col1/
col3 further down) — stop scanning at Grand Total or it gets misread.

Each workbook may have a second "Derivative...` sheet (SEBI derivative-
disclosure notice, not a holdings table) — it has no "Name Of the
Instrument" header row, so parse_sheet() naturally returns no fund_name for
it and callers skip it, same as any sheet that doesn't match.
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

try:
    import xlrd
except ImportError:  # only needed for the rare legacy-.xls file (see load_workbook_sheets)
    xlrd = None

ROOT = Path(__file__).parent
CACHE_DIR = ROOT / "cache" / "hdfc" / "xlsx"
OUT_DIR = ROOT / "out" / "hdfc_xlsx"

SKIP_RE = re.compile(r"^(sub\s*total|total|grand\s*total)\b", re.IGNORECASE)

COL_ISIN = 1
COL_COUPON = 2
COL_NAME = 3
COL_INDUSTRY = 4
COL_QUANTITY = 5
COL_MARKET_VALUE = 6
COL_WEIGHT = 7

SECTION_TYPE_MAP = [
    (re.compile(r"certificate of deposit", re.I), "cd"),
    (re.compile(r"commercial paper", re.I), "cp"),
    (re.compile(r"treasury\s*bill|t-?\s*bills?\b", re.I), "tbill"),
    (re.compile(r"government (bond|security|securities)|g-?sec|state (govern|gover)ment|state development loan", re.I), "gsec"),
    (re.compile(r"treps|reverse repo|repo", re.I), "treps"),
    (re.compile(r"mutual fund", re.I), "fund"),
    (
        re.compile(
            r"corporate (debt|bond)|debenture|\bncd\b|non.convertible|money market|"
            r"\bdebt instrument|\bdebt securit",
            re.I,
        ),
        "corporate_debt",
    ),
    (re.compile(r"cash|net current|net receivable", re.I), "cash"),
    (re.compile(r"reit|invit", re.I), "reit"),
    (re.compile(r"future|option|derivative|hedg", re.I), "derivative"),
    (re.compile(r"preference share", re.I), "preference"),
    (re.compile(r"equity", re.I), "equity"),
]


def classify(section_label: str, name: str = "") -> str:
    for text in (section_label, name):
        for pat, kind in SECTION_TYPE_MAP:
            if pat.search(text or ""):
                return kind
    return "equity"


def num(v):
    if v is None or v == "" or v == "NIL":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def load_workbook_sheets(path: Path):
    """Every sample inspected during initial development was real OOXML, but at
    least one confirmed exception exists (2026-01 HDFC Retirement Savings Fund -
    Hybrid-Debt Plan came back as a genuine legacy BIFF8/OLE2 compound file despite
    the ".xlsx" URL/filename — same "extension lies" edge case PPFAS's format has,
    see parse_ppfas_xlsx.py). Detect by file signature, not extension, same fix."""
    with open(path, "rb") as f:
        sig = f.read(8)
    if sig[:4] == b"PK\x03\x04":
        wb = openpyxl.load_workbook(str(path), data_only=True)
        return {name: [list(row) for row in wb[name].iter_rows(values_only=True)] for name in wb.sheetnames}
    if sig[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        if xlrd is None:
            raise RuntimeError(f"{path} is legacy BIFF8/OLE2 but xlrd isn't installed")
        wb = xlrd.open_workbook(str(path))
        return {name: [wb.sheet_by_name(name).row_values(i) for i in range(wb.sheet_by_name(name).nrows)]
                for name in wb.sheet_names()}
    raise RuntimeError(f"{path}: unrecognized file signature {sig!r}")


def find_fund_name(rows):
    """Row 0, col 0 is always the scheme's full descriptive name followed by
    a parenthetical scheme-type description — e.g. "HDFC NIFTY 50 ETF (An
    open ended scheme replicating / tracking NIFTY 50 Index)". Unlike PPFAS,
    can't key off the word "fund" (many ETF/index titles don't contain it),
    so just take row0/col0 directly and strip the trailing parenthetical."""
    if not rows or not rows[0]:
        return None
    cell = rows[0][0] if len(rows[0]) > 0 else None
    if not isinstance(cell, str) or not cell.strip():
        return None
    # Description is usually "(...)" but at least one confirmed month (Feb-Apr 2026,
    # HDFC Nifty India Consumption Index Fund) uses "[...]" instead — strip either.
    return re.sub(r"\s*[\(\[].*$", "", cell).strip()


def cell(r, idx):
    return r[idx] if idx is not None and idx < len(r) else None


def parse_sheet(rows):
    header_idx = next((i for i, r in enumerate(rows) if r and any(
        isinstance(c, str) and re.search(r"name of the instrument", c, re.I) for c in r
    )), None)
    if header_idx is None:
        return {"holdings": [], "grand_total": None, "fund_name": None}

    fund_name = find_fund_name(rows)
    if not fund_name:
        return {"holdings": [], "grand_total": None, "fund_name": None}

    holdings = []
    section_label = None
    grand_total = None

    for r in rows[header_idx + 1:]:
        if not r:
            continue
        raw_name = cell(r, COL_NAME)
        name = raw_name.strip() if isinstance(raw_name, str) else raw_name
        label_cell = cell(r, COL_ISIN)
        label = label_cell.strip() if isinstance(label_cell, str) else None

        # A pure section-marker / Sub Total / Total / Grand Total row has its text in
        # col1 (the ISIN column doubling as the label column) and no Name in col3.
        if label and not name:
            if SKIP_RE.match(label):
                if label.upper() == "GRAND TOTAL":
                    grand_total = num(cell(r, COL_WEIGHT))
                    # Everything after Grand Total is a different table (Top Ten
                    # Holdings / NAV history / riskometer) with numeric junk in the
                    # same column positions — stop here, same fix as PPFAS's.
                    break
                continue
            # Only overwrite the tracked label when the new text itself matches a
            # known category — nested structural sub-headers ("(a) Listed /
            # awaiting listing...") are reused verbatim across sections and carry
            # no type information of their own (same bug PPFAS had to avoid).
            if any(pat.search(label) for pat, _ in SECTION_TYPE_MAP):
                section_label = label
            continue

        if not name:
            continue
        if isinstance(name, str) and SKIP_RE.match(name):
            continue

        weight_frac = num(cell(r, COL_WEIGHT))
        if weight_frac is None:
            continue

        isin = label if label else ""
        industry = cell(r, COL_INDUSTRY)
        industry = industry.strip() if isinstance(industry, str) else (industry or "")
        quantity = num(cell(r, COL_QUANTITY))
        market_value_lacs = num(cell(r, COL_MARKET_VALUE))
        holdings.append({
            "name": name.strip() if isinstance(name, str) else str(name),
            "isin": isin,
            "industry": industry,
            "quantity": quantity,
            # Lacs -> Crores (HDFC discloses in Lacs; PPFAS's source is already Crores).
            "market_value_cr": round(market_value_lacs / 100, 4) if market_value_lacs is not None else None,
            # Unlike PPFAS's source (a fraction, *100 needed), HDFC's "% to NAV"
            # column is already a percentage (e.g. 92.75, Grand Total = 100) —
            # confirmed by direct inspection, no scaling needed.
            "weight": round(weight_frac, 4),
            "type": classify(section_label, name if isinstance(name, str) else ""),
        })

    return {"holdings": holdings, "grand_total": grand_total, "fund_name": fund_name}


def main():
    period = sys.argv[1] if len(sys.argv) > 1 else None
    xlsx_path = Path(sys.argv[2]) if len(sys.argv) > 2 else None
    if not xlsx_path:
        print("usage: parse_hdfc_xlsx.py <period YYYY-MM> <path to .xlsx>")
        return 1

    sheets = load_workbook_sheets(xlsx_path)
    funds = {}
    for sheet_name, rows in sheets.items():
        parsed = parse_sheet(rows)
        if not parsed["fund_name"]:
            continue
        total = round(sum(h["weight"] for h in parsed["holdings"]), 2)
        gt = parsed["grand_total"]
        gt_pct = round(gt, 2) if gt is not None else None
        ok = gt_pct is not None and abs(total - gt_pct) < 0.5
        print(f"  {parsed['fund_name']}: {len(parsed['holdings'])} holdings, "
              f"sum={total:.2f}%, GRAND TOTAL={gt_pct}%, {'OK' if ok else 'MISMATCH'}")
        funds[parsed["fund_name"]] = {"holdings": parsed["holdings"], "grand_total_pct": gt_pct}

    if period:
        OUT_DIR.mkdir(parents=True, exist_ok=True)
        out_path = OUT_DIR / f"{period}.json"
        out_path.write_text(json.dumps(funds, indent=2))
        print(f"Written to {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
