#!/usr/bin/env python3
"""Deterministic parser for Mirae Asset Mutual Fund's "Detailed Portfolio
Disclosure" XLSX — the SEBI-mandated monthly portfolio statement, NOT the
marketing factsheet PDF at /downloads/factsheet (see discover_mirae_xlsx.py's
docstring and lib/types.ts's header comment: the factsheet PDF has no
ISIN/quantity per holding, the XLSX does).

Layout and row-classification rules are structurally identical to PPFAS's
(see parse_ppfas_xlsx.py, whose docstring this mirrors) — confirmed by
downloading and inspecting real files directly: same column set (Name of the
Instrument | ISIN | Industry ^/ Rating | Quantity | Market/Fair Value (Rs. in
Lacs) | % to Net Assets | YTM), same "weight-presence = data row" rule, same
Sub Total/Total-skip-but-GRAND-TOTAL-anchors-trust convention, same nested
section-header classify() fix. Two file shapes share this exact per-sheet
layout, both handled by parse_sheet() below:
  - 2022-present: one XLSX per scheme per month, single sheet.
  - up to 2021: one combined workbook per month, one sheet per scheme (like
    PPFAS's own workbook shape) plus a SUMMARY sheet (skipped — no "Name of
    the Instrument" header row, so parse_sheet() naturally returns empty for
    it and main() skips schemeless results).

Fund-name extraction differs from PPFAS's "longest cell containing 'Fund'"
heuristic — that heuristic picks the WRONG cell here, because row2 is a
parenthetical scheme description that's usually longer than the actual name
in row0/row4 and also contains "Fund" (e.g. row0 "Mirae Asset Flexi Cap
Fund", row2 "(Flexi Cap Fund - An open ended dynamic equity scheme investing
across large cap, mid cap, small cap stocks)"). Confirmed by direct
inspection of multiple scheme files (equity, debt, index, ETF) — the clean
name is reliably the first "Mirae Asset ..." cell that does NOT start with
"(", found by scanning top-to-bottom instead of longest-match.

Also extracts optional PortfolioMetrics (YTM / Macaulay duration / residual
maturity — present for debt-oriented schemes, absent/blank for pure equity,
both confirmed by inspection) from labeled rows near the sheet's tail, since
lib/types.ts's AnalyseData.metrics field already exists for this.
"""
import json
import re
import sys
from pathlib import Path

import openpyxl
import xlrd

ROOT = Path(__file__).parent
CACHE_DIR = ROOT / "cache" / "mirae" / "xlsx"
OUT_DIR = ROOT / "out" / "mirae_xlsx"

SKIP_RE = re.compile(r"^(sub\s*total|total|grand\s*total)\b", re.IGNORECASE)

SECTION_TYPE_MAP = [
    (re.compile(r"certificate of deposit", re.I), "cd"),
    (re.compile(r"commercial paper", re.I), "cp"),
    (re.compile(r"treasury\s*bill|t-?\s*bills?\b", re.I), "tbill"),
    (re.compile(r"government (bond|security|securities)|g-?sec|state (govern|gover)ment|state development loan", re.I), "gsec"),
    (re.compile(r"treps|reverse repo|repo", re.I), "treps"),
    (re.compile(r"mutual fund", re.I), "fund"),
    (
        re.compile(
            r"corporate (debt|bond)|debenture|\bncd\b|non.convertible|money market|"
            r"\bdebt instrument|\bdebt securit",
            re.I,
        ),
        "corporate_debt",
    ),
    (re.compile(r"cash|net (current|receivable)", re.I), "cash"),
    (re.compile(r"reit|invit", re.I), "reit"),
    (re.compile(r"future|option|derivative|hedg", re.I), "derivative"),
    (re.compile(r"equity", re.I), "equity"),
]


def classify(section_label: str, name: str = "") -> str:
    for text in (section_label, name):
        for pat, kind in SECTION_TYPE_MAP:
            if pat.search(text or ""):
                return kind
    return "equity"


def num(v):
    if v is None or v == "" or v == "NIL":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


COLUMN_NAME_PATTERNS = {
    "name": re.compile(r"name of the instrument", re.I),
    "isin": re.compile(r"^isin", re.I),
    "industry": re.compile(r"industry|rating", re.I),
    "quantity": re.compile(r"quantity", re.I),
    "market_value": re.compile(r"market.*value", re.I),
    "weight": re.compile(r"%\s*to\s*(net|aum)", re.I),
}


def load_workbook_rows(path: Path):
    """Same real-OOXML-vs-legacy-BIFF8 detection as parse_ppfas_xlsx.py's
    load_workbook_rows — Mirae's archive serves both under .xls/.xlsx
    extensions inconsistently across eras, confirmed by inspecting file
    signatures directly (pre-2019 files are genuine BIFF8)."""
    with open(path, "rb") as f:
        sig = f.read(4)
    if sig == b"PK\x03\x04":
        wb = openpyxl.load_workbook(str(path), data_only=True)
        return {name: [list(row) for row in wb[name].iter_rows(values_only=True)] for name in wb.sheetnames}
    wb = xlrd.open_workbook(str(path))
    return {name: [wb.sheet_by_name(name).row_values(i) for i in range(wb.sheet_by_name(name).nrows)]
            for name in wb.sheet_names()}


TRAILING_PAREN_RE = re.compile(r"\s*\([^()]*\)\s*$")


def find_fund_name(rows):
    """First 'Mirae Asset ...' cell in the first ~8 rows that isn't a
    parenthetical description — see module docstring for why PPFAS's
    longest-match heuristic doesn't work here.

    Strips a trailing parenthetical suffix ("... (Formerly Known as ...)",
    "... (MAGOLDETF)") off the raw cell — confirmed by inspection this suffix
    is baked directly into row0's fund-name cell for schemes renamed within
    the archive window, and left unstripped it fragments one scheme's history
    into two different canonical names (pre-/post-rename) instead of joining
    them under the identity build_dataset_mirae.py resolves against AMFI."""
    for r in rows[:8]:
        for c in r:
            if not isinstance(c, str):
                continue
            t = c.strip()
            if t.startswith("(") or len(t) <= 10:
                continue
            if re.search(r"mirae\s*asset", t, re.I):
                return TRAILING_PAREN_RE.sub("", t).strip()
    return None


def locate_columns(header_row):
    cols = {}
    for i, cell in enumerate(header_row):
        if not isinstance(cell, str):
            continue
        for key, pat in COLUMN_NAME_PATTERNS.items():
            if key not in cols and pat.search(cell):
                cols[key] = i
    return cols


METRIC_LABELS = {
    "ytm": re.compile(r"annualised portfolio ytm", re.I),
    "macaulay_days": re.compile(r"macaulay duration", re.I),
    "residual_days": re.compile(r"residual maturity", re.I),
}


def find_metrics(rows):
    """Best-effort scan for the labeled YTM/duration rows near a debt
    scheme's tail (see module docstring — absent for pure equity schemes,
    which is fine, all fields are optional)."""
    metrics = {"ytm": None, "macaulay_days": None, "residual_days": None,
               "benchmark": None, "inception": None, "fund_managers": None}
    for r in rows:
        if not r:
            continue
        for key, pat in METRIC_LABELS.items():
            if metrics[key] is not None:
                continue
            label = next((c for c in r if isinstance(c, str)), None)
            if label and pat.search(label):
                for c in r:
                    v = num(c)
                    if v is not None:
                        metrics[key] = v
                        break
    if metrics["ytm"] is not None:
        metrics["ytm"] = round(metrics["ytm"] * 100, 4)  # fraction -> %
    if not any(metrics.values()):
        return None
    return metrics


def parse_sheet(rows):
    header_idx = next((i for i, r in enumerate(rows) if r and any(
        isinstance(c, str) and COLUMN_NAME_PATTERNS["name"].search(c) for c in r
    )), None)
    if header_idx is None:
        return {"holdings": [], "grand_total": None, "fund_name": None, "metrics": None}

    cols = locate_columns(rows[header_idx])
    if "name" not in cols or "weight" not in cols:
        return {"holdings": [], "grand_total": None, "fund_name": None, "metrics": None}

    fund_name = find_fund_name(rows)
    if not fund_name:
        return {"holdings": [], "grand_total": None, "fund_name": None, "metrics": None}

    holdings = []
    section_label = None
    grand_total = None
    tail_start = None

    def cell(r, key):
        idx = cols.get(key)
        return r[idx] if idx is not None and idx < len(r) else None

    for offset, r in enumerate(rows[header_idx + 1:]):
        if not r:
            continue
        raw_name = cell(r, "name")
        name = raw_name.strip() if isinstance(raw_name, str) else raw_name
        if not name:
            continue
        if isinstance(name, str) and SKIP_RE.match(name):
            if name.strip().upper() == "GRAND TOTAL":
                grand_total = num(cell(r, "weight"))
                tail_start = header_idx + 1 + offset
                break
            continue

        weight_frac = num(cell(r, "weight"))
        if weight_frac is None:
            if isinstance(name, str) and any(pat.search(name) for pat, _ in SECTION_TYPE_MAP):
                section_label = name
            continue

        isin = cell(r, "isin")
        industry = cell(r, "industry")
        industry = industry.strip() if isinstance(industry, str) else (industry or "")
        quantity = num(cell(r, "quantity"))
        market_value_lakhs = num(cell(r, "market_value"))
        holdings.append({
            "name": name.strip() if isinstance(name, str) else str(name),
            "isin": (isin.strip() if isinstance(isin, str) else "") or "",
            "industry": industry,
            "quantity": quantity,
            "market_value_cr": round(market_value_lakhs / 100, 4) if market_value_lakhs is not None else None,
            "weight": round(weight_frac * 100, 4),
            "type": classify(section_label, name if isinstance(name, str) else ""),
        })

    metrics = find_metrics(rows[tail_start:] if tail_start else rows[header_idx:])
    return {"holdings": holdings, "grand_total": grand_total, "fund_name": fund_name, "metrics": metrics}


def main():
    period = sys.argv[1] if len(sys.argv) > 1 else None
    xlsx_path = Path(sys.argv[2]) if len(sys.argv) > 2 else None
    if not xlsx_path:
        print("usage: parse_mirae_xlsx.py <period YYYY-MM> <path to .xlsx>")
        return 1

    sheets = load_workbook_rows(xlsx_path)
    funds = {}
    for sheet_name, rows in sheets.items():
        parsed = parse_sheet(rows)
        if not parsed["fund_name"]:
            continue
        total = round(sum(h["weight"] for h in parsed["holdings"]), 2)
        gt = parsed["grand_total"]
        gt_pct = round(gt * 100, 2) if gt is not None else None
        ok = gt_pct is not None and abs(total - gt_pct) < 0.5
        print(f"  {parsed['fund_name']}: {len(parsed['holdings'])} holdings, "
              f"sum={total:.2f}%, GRAND TOTAL={gt_pct}%, {'OK' if ok else 'MISMATCH'}")
        funds[parsed["fund_name"]] = {
            "holdings": parsed["holdings"], "grand_total_pct": gt_pct, "metrics": parsed["metrics"],
        }

    if period:
        OUT_DIR.mkdir(parents=True, exist_ok=True)
        out_path = OUT_DIR / f"{period}.json"
        existing = {}
        if out_path.exists():
            existing = json.loads(out_path.read_text())
        existing.update(funds)
        out_path.write_text(json.dumps(existing, indent=2))
        print(f"Written to {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
