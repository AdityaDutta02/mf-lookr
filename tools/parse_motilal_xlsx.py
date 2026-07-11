#!/usr/bin/env python3
"""Deterministic parser for Motilal Oswal MF's "Scheme Portfolio Details" /
"Month End Portfolio" XLS/XLSX — the SEBI-mandated monthly full portfolio
disclosure, NOT the marketing factsheet PDF (see discover_motilal_xlsx.py's
docstring for the PDF-vs-XLS distinction that matters here).

Structurally very close to PPFAS's format (tools/parse_ppfas_xlsx.py) — same
skip/classify approach — with three real differences:

1. ONE WORKBOOK PER MONTH, MANY SCHEMES PER WORKBOOK. Newer files (~2023+)
   also carry an "Index" sheet (Sr No | Fund Name | Fund Code) mapping sheet
   codes like "YO01" -> "Motilal Oswal Nifty 50 ETF" — but that mapping has at
   least one observed typo (index says "Y054", the actual sheet is "YO54") and
   doesn't exist at all pre-2023, so it is NOT used for identity. Instead each
   sheet's own title block is read directly (see find_fund_name below) — this
   works uniformly across every era.

2. WEIGHT IS ALREADY A PERCENTAGE (e.g. 11.14), NOT A FRACTION (e.g. 0.1114)
   the way PPFAS's "% to Net Assets" column is. No *100 anywhere here.

3. LEGACY ERA (pre ~2022) FILES ARE GENUINE BIFF8 .xls (not renamed OOXML like
   PPFAS's old files) — load_workbook_rows() below still dispatches on the
   actual file signature rather than trusting the extension, for the same
   reason PPFAS's does.

Row classification reuses PPFAS's core insight: a row with a populated weight
cell is a holding; a row with none is a section-label row that updates
classification context for holdings beneath it, and only overwrites the
tracked label when the label text itself matches a known instrument-type
pattern (so nested sub-headers like "(a) Listed / awaiting listing..." can't
wipe out a correct parent label like "Certificate of Deposit").
"""
import json
import re
import sys
from pathlib import Path

import openpyxl
import xlrd

ROOT = Path(__file__).parent
CACHE_DIR = ROOT / "cache" / "motilal" / "xlsx"
OUT_DIR = ROOT / "out" / "motilal_xlsx"

SKIP_RE = re.compile(r"^(sub\s*total|total|grand\s*total)\b", re.IGNORECASE)

SECTION_TYPE_MAP = [
    # Checked before the generic corporate_debt pattern below — "Corporate Debt
    # Market Development Fund" (the SEBI-mandated CDMDF allocation, present as
    # its own tiny section in every debt-oriented sheet since ~2024) would
    # otherwise false-positive-match "corporate debt".
    (re.compile(r"development fund", re.I), "fund"),
    (re.compile(r"certificate of deposit", re.I), "cd"),
    (re.compile(r"commercial paper", re.I), "cp"),
    (re.compile(r"treasury\s*bill|t-?\s*bills?\b|cash management bill", re.I), "tbill"),
    (re.compile(r"government (bond|security|securities)|g-?sec|state (govern|gover)ment|state development loan", re.I), "gsec"),
    (re.compile(r"treps|cblo|reverse repo|repo", re.I), "treps"),
    (re.compile(r"mutual fund|fund of fund", re.I), "fund"),
    (
        re.compile(
            r"corporate (debt|bond)|debenture|\bncds?\b|non.convertible|money market|"
            r"\bdebt instrument|\bdebt securit|\bbonds?\b",
            re.I,
        ),
        "corporate_debt",
    ),
    (re.compile(r"cash|net receivable", re.I), "cash"),
    (re.compile(r"reit|invit", re.I), "reit"),
    (re.compile(r"future|option|derivative|hedg", re.I), "derivative"),
    (re.compile(r"equity", re.I), "equity"),
]


def classify(section_label: str, name: str = "") -> str:
    for text in (section_label, name):
        for pat, kind in SECTION_TYPE_MAP:
            if pat.search(text or ""):
                return kind
    return "equity"


def num(v):
    if v is None or v == "" or (isinstance(v, str) and v.strip().upper() in ("", "NIL", "-")):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


COLUMN_NAME_PATTERNS = {
    "name": re.compile(r"name of (the )?instrument", re.I),
    "isin": re.compile(r"^isin", re.I),
    "industry": re.compile(r"industry|rating", re.I),
    "quantity": re.compile(r"quantity", re.I),
    "market_value": re.compile(r"market.*value", re.I),
    "weight": re.compile(r"%\s*to\s*(net|aum|nav)", re.I),
}

# The genuine per-sheet scheme name always starts with "Motilal Oswal" or, for a
# handful of newer funds, the abbreviated "MO " (e.g. "MO Nifty Capital Market
# Index Fund"). Boilerplate header lines that also contain "Motilal Oswal" (AMC
# legal name, "Investment Manager for..." disclaimer) are excluded by keyword.
FUND_NAME_START_RE = re.compile(r"^(Motilal Oswal|MO )")
FUND_NAME_EXCLUDE_RE = re.compile(
    r"asset management company|registered office|investment manager|"
    r"monthly portfolio|portfolio statement|portfolio as on|statement as on|mutual fund$",
    re.I,
)


def load_workbook_rows(path: Path):
    """Returns {sheet_name: [[cell, ...], ...]} — real OOXML (openpyxl) or
    legacy BIFF8 (xlrd), detected by file signature, not extension (see
    module docstring point 3)."""
    with open(path, "rb") as f:
        sig = f.read(4)
    if sig == b"PK\x03\x04":
        wb = openpyxl.load_workbook(str(path), data_only=True)
        return {name: [list(row) for row in wb[name].iter_rows(values_only=True)] for name in wb.sheetnames}
    wb = xlrd.open_workbook(str(path))
    return {name: [wb.sheet_by_name(name).row_values(i) for i in range(wb.sheet_by_name(name).nrows)]
            for name in wb.sheet_names()}


def find_fund_name(rows):
    for r in rows[:12]:
        for c in r:
            if not isinstance(c, str):
                continue
            s = c.strip()
            if s and FUND_NAME_START_RE.match(s) and not FUND_NAME_EXCLUDE_RE.search(s):
                return re.sub(r"\s*\(.*$", "", s).strip()
    return None


def locate_columns(header_row):
    cols = {}
    for i, cell in enumerate(header_row):
        if not isinstance(cell, str):
            continue
        for key, pat in COLUMN_NAME_PATTERNS.items():
            if key not in cols and pat.search(cell):
                cols[key] = i
    return cols


def parse_sheet(rows):
    header_idx = next((i for i, r in enumerate(rows) if r and any(
        isinstance(c, str) and COLUMN_NAME_PATTERNS["name"].search(c) for c in r
    )), None)
    if header_idx is None:
        return {"holdings": [], "grand_total": None, "fund_name": None}

    cols = locate_columns(rows[header_idx])
    if "name" not in cols or "weight" not in cols:
        return {"holdings": [], "grand_total": None, "fund_name": None}

    fund_name = find_fund_name(rows)
    if not fund_name:
        return {"holdings": [], "grand_total": None, "fund_name": None}

    holdings = []
    section_label = None
    grand_total = None

    def cell(r, key):
        idx = cols.get(key)
        return r[idx] if idx is not None and idx < len(r) else None

    for r in rows[header_idx + 1:]:
        if not r:
            continue
        raw_name = cell(r, "name")
        name = raw_name.strip() if isinstance(raw_name, str) else raw_name
        raw_isin = cell(r, "isin")
        isin_val = (raw_isin.strip() if isinstance(raw_isin, str) else raw_isin) or ""
        # Some Motilal "Mutual Fund Units" sub-sections (fund-of-funds / passive
        # FOF schemes holding other Motilal funds) ship with a genuinely BLANK
        # instrument-name cell — only ISIN + weight are populated. Dropping
        # those rows for "no name" silently loses real weight (seen: FOF
        # schemes summing to ~62% instead of 100%). Only truly empty rows (no
        # name AND no ISIN) get skipped; a blank name with an ISIN falls back
        # to the ISIN as its display name below instead of being discarded.
        if not name and not isin_val:
            continue
        if isinstance(name, str) and SKIP_RE.match(name):
            if name.strip().upper() == "GRAND TOTAL":
                grand_total = num(cell(r, "weight"))
                # Everything after GRAND TOTAL is TER / NAV-history / notes —
                # different table, same column positions, would misread as junk
                # holdings if we kept going.
                break
            continue

        weight = num(cell(r, "weight"))
        if weight is None:
            if isinstance(name, str) and any(pat.search(name) for pat, _ in SECTION_TYPE_MAP):
                section_label = name
            continue

        industry = cell(r, "industry")
        industry = industry.strip() if isinstance(industry, str) else (industry or "")
        quantity = num(cell(r, "quantity"))
        market_value_lakhs = num(cell(r, "market_value"))
        holdings.append({
            "name": (name.strip() if isinstance(name, str) else str(name)) or isin_val or "Unnamed instrument",
            "isin": isin_val,
            "industry": industry,
            "quantity": quantity,
            "market_value_cr": round(market_value_lakhs / 100, 4) if market_value_lakhs is not None else None,
            "weight": round(weight, 4),  # already a percentage — see module docstring point 2
            "type": classify(section_label, name if isinstance(name, str) else ""),
        })

    # Weight is USUALLY already a percentage (see module docstring point 2), but
    # this is not consistent across Motilal's own 13+ years of filings — several
    # eras (observed: most of 2021 through March 2026) ship "% to Net Assets" as
    # a genuine fraction instead (e.g. 0.1114, GRAND TOTAL 1.0), like PPFAS's
    # format. Detect per-sheet from the GRAND TOTAL magnitude (or the holdings'
    # own weight sum when GRAND TOTAL is missing) rather than trusting a fixed
    # era cutoff, and normalize to percent so downstream consumers never have to
    # care which era a given month came from.
    reference = grand_total if grand_total is not None else (
        sum(h["weight"] for h in holdings) if holdings else None
    )
    if reference is not None and 0 < reference <= 5:
        for h in holdings:
            h["weight"] = round(h["weight"] * 100, 4)
        grand_total = round(grand_total * 100, 4) if grand_total is not None else None

    return {"holdings": holdings, "grand_total": grand_total, "fund_name": fund_name}


def main():
    period = sys.argv[1] if len(sys.argv) > 1 else None
    xlsx_path = Path(sys.argv[2]) if len(sys.argv) > 2 else None
    if not xlsx_path:
        print("usage: parse_motilal_xlsx.py <period YYYY-MM> <path to .xlsx/.xls>")
        return 1

    sheets = load_workbook_rows(xlsx_path)
    funds = {}
    for sheet_name, rows in sheets.items():
        if sheet_name.strip().lower() == "index":
            continue
        parsed = parse_sheet(rows)
        if not parsed["fund_name"]:
            continue
        total = round(sum(h["weight"] for h in parsed["holdings"]), 2)
        gt_pct = round(parsed["grand_total"], 2) if parsed["grand_total"] is not None else None
        ok = gt_pct is not None and abs(total - gt_pct) < 0.5
        print(f"  {parsed['fund_name']}: {len(parsed['holdings'])} holdings, "
              f"sum={total:.2f}%, GRAND TOTAL={gt_pct}%, {'OK' if ok else 'MISMATCH'}")
        funds[parsed["fund_name"]] = {"holdings": parsed["holdings"], "grand_total_pct": gt_pct}

    if period:
        OUT_DIR.mkdir(parents=True, exist_ok=True)
        out_path = OUT_DIR / f"{period}.json"
        out_path.write_text(json.dumps(funds, indent=2))
        print(f"Written to {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
