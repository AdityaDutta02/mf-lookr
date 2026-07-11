#!/usr/bin/env python3
"""Deterministic parser for Invesco Mutual Fund's "Monthly Portfolio
Statement" xlsx — the SEBI-mandated detailed disclosure (ISIN + quantity per
holding), NOT the marketing factsheet PDF under the site's "Factsheets" tab.
Confirmed by direct inspection of samples across every classification the
site offers (equity, fixed-income/debt, hybrid, ETF, index fund, FoF): one
file per scheme per month (same shape as HDFC's archive, unlike PPFAS's
one-workbook-many-sheets), first sheet = holdings, second sheet ("RTP" /
Top 10 Holdings) = a different table (no "Name of the Instrument" header),
naturally skipped by the same header-not-found check every other AMC's
parser uses.

Layout — like PPFAS (NOT HDFC's fixed-column layout), because column COUNT
varies by scheme type: a Fund-of-Funds sheet drops the "Industry*/Rating"
column entirely (7 columns instead of 8), so columns are located by header
TEXT via regex, never by fixed index:
  Name of the Instrument | ISIN | [Industry*/Rating] | Quantity |
  Market/Fair Value (Rs. in Lakhs) | % to Net Assets | YTM

Row classification: identical "weight presence, not ISIN presence" rule as
PPFAS/HDFC — a row with a non-empty "% to Net Assets" cell is a real holding
(TREPS/Net Receivables have no ISIN but do have a weight); a row with a
label in the Name column and no weight is a section header, tracked via the
same nested-label fix (only overwrite when the label ITSELF matches a known
category — "(a) Listed / awaiting listing..." is reused across sections and
carries no type info of its own).

Units: unlike PPFAS's "% to Net Assets" (a 0-1 fraction) but LIKE HDFC's
"% to NAV", Invesco's weight column is already a percentage (e.g. 4.08, not
0.0408; GRAND TOTAL = 100) — confirmed by direct inspection, no *100 needed.
Market value is "Rs. in Lakhs" like both PPFAS and HDFC — divide by 100 for
the app's ₹cr unit.

File-signature quirk: some pre-~2020 files are served under a ".xls" URL
but are real OOXML zips — SAME "extension lies" bug PPFAS's and HDFC's
archives have. Worse here: openpyxl's own path-based loader refuses ANY
filename ending ".xls" regardless of actual content (its _validate_archive
hard-codes that extension check) — so we always load via an in-memory
BytesIO buffer instead of a path string, which skips that check entirely
and reads the real bytes. A handful of genuinely legacy BIFF8 files may
still exist (not confirmed for Invesco, but PPFAS/HDFC both have some) —
detect by file signature and fall back to xlrd, same as those two parsers.
"""
import io
import json
import re
import sys
from pathlib import Path

import openpyxl

try:
    import xlrd
except ImportError:  # only needed for a genuine legacy BIFF8 file, if one turns up
    xlrd = None

ROOT = Path(__file__).parent
CACHE_DIR = ROOT / "cache" / "invesco" / "xlsx"
OUT_DIR = ROOT / "out" / "invesco_xlsx"

SKIP_RE = re.compile(r"^(sub\s*total|total|grand\s*total)\b", re.IGNORECASE)

SECTION_TYPE_MAP = [
    (re.compile(r"certificate of deposit", re.I), "cd"),
    (re.compile(r"commercial paper", re.I), "cp"),
    (re.compile(r"treasury\s*bill|t-?\s*bills?\b", re.I), "tbill"),
    (re.compile(r"government (bond|security|securities)|g-?sec|state (govern|gover)ment|state development loan", re.I), "gsec"),
    (re.compile(r"treps|reverse repo|repo", re.I), "treps"),
    (re.compile(r"exchange traded fund|mutual fund", re.I), "fund"),
    (
        re.compile(
            r"corporate (debt|bond)|debenture|\bncd\b|non.convertible|money market|"
            r"\bdebt instrument|\bdebt securit",
            re.I,
        ),
        "corporate_debt",
    ),
    (re.compile(r"cash|net current|net receivable", re.I), "cash"),
    (re.compile(r"reit|invit", re.I), "reit"),
    (re.compile(r"future|option|derivative|hedg", re.I), "derivative"),
    (re.compile(r"preference share", re.I), "preference"),
    (re.compile(r"equity", re.I), "equity"),
]


def classify(section_label: str, name: str = "") -> str:
    for text in (section_label, name):
        for pat, kind in SECTION_TYPE_MAP:
            if pat.search(text or ""):
                return kind
    return "equity"


NUM_PREFIX_RE = re.compile(r"^\s*(-?[\d,]+\.?\d*)")


def num(v):
    if v is None or v == "" or v == "NIL" or v == "Nil":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        # Some weight/quantity cells carry a trailing footnote marker glued
        # onto the number as a string, e.g. "1.41 $" for lock-in shares
        # ("$ Values post illiquidity discount on account of lock-in
        # period" — a confirmed 2021 quirk) — strip it rather than drop the
        # whole holding, which otherwise silently under-counts the weight
        # total (confirmed: ESG Equity Fund / Focused 20 Equity Fund
        # 2021-08..2022-08 trust-band mismatches were exactly this, off by
        # the missing lock-in-share row's weight to the cent).
        if isinstance(v, str):
            m = NUM_PREFIX_RE.match(v)
            if m and m.group(1) not in ("", "-", "."):
                try:
                    return float(m.group(1).replace(",", ""))
                except ValueError:
                    return None
        return None


COLUMN_NAME_PATTERNS = {
    "name": re.compile(r"name of the instrument", re.I),
    "isin": re.compile(r"^isin", re.I),
    "industry": re.compile(r"industry|rating", re.I),
    "quantity": re.compile(r"quantity", re.I),
    "market_value": re.compile(r"market.*value", re.I),
    "weight": re.compile(r"%\s*to\s*(net\s*assets|nav|aum)", re.I),
}


def load_workbook_sheets(path: Path):
    """Always load via an in-memory buffer, not a path string — openpyxl's own
    path-based loader hard-refuses any ".xls"-named file regardless of actual
    content (see module docstring), and a BytesIO buffer skips that
    extension check entirely. Detect real legacy BIFF8 by signature and fall
    back to xlrd for that (rare) case."""
    raw = path.read_bytes()
    if raw[:4] == b"PK\x03\x04":
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        return {name: [list(row) for row in wb[name].iter_rows(values_only=True)] for name in wb.sheetnames}
    if raw[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        if xlrd is None:
            raise RuntimeError(f"{path} is legacy BIFF8/OLE2 but xlrd isn't installed")
        wb = xlrd.open_workbook(file_contents=raw)
        return {name: [wb.sheet_by_name(name).row_values(i) for i in range(wb.sheet_by_name(name).nrows)]
                for name in wb.sheet_names()}
    raise RuntimeError(f"{path}: unrecognized file signature {raw[:8]!r}")


STATEMENT_LABEL_RE = re.compile(r"monthly portfolio statement", re.I)


def find_fund_name(rows):
    """Structural, not content-keyword-based: every sample inspected (recent
    and pre-2020 alike) has a "Monthly Portfolio Statement as on <date>" row
    adjacent to a row whose first text cell is the scheme's full name + a
    parenthetical scheme-type description. PPFAS's/HDFC's "cell contains the
    word 'fund'" heuristic doesn't work here — at least one confirmed older
    scheme name ("Invesco India Tax Plan", pre-rename to "...ELSS Tax Saver
    Fund") has no literal "fund" substring at all — so key off the
    statement-label row's position instead.

    Usually the name row comes right AFTER the statement-label row (name +
    parenthetical description on one line), but two other layouts are
    confirmed in the wild:
      - 2024-03 (all 37 schemes): opposite order — name row immediately
        BEFORE the statement-label row, with a blank row after it.
      - some other pre-2025 files (e.g. "Invesco India PSU Equity Fund",
        2024-03): the name and its parenthetical description are on TWO
        separate rows, both BEFORE the statement-label row (name, then a
        row that is ONLY the parenthetical) — stripping the parenthetical
        from that continuation row alone leaves "", so it must be skipped
        in favor of the row before it.
    So: scan outward from the statement-label row (immediate neighbors
    first), skip any candidate row that is nothing but a parenthetical
    continuation, and return the first candidate that still has real text
    after stripping a trailing parenthetical."""
    for i, r in enumerate(rows[:8]):
        if any(isinstance(c, str) and STATEMENT_LABEL_RE.search(c) for c in r):
            for j in (i + 1, i - 1, i - 2, i + 2, i - 3, i + 3):
                if j == i or not (0 <= j < len(rows)):
                    continue
                for c in rows[j]:
                    if not (isinstance(c, str) and c.strip()):
                        continue
                    text = c.strip()
                    if text.startswith("(") or text.startswith("["):
                        continue  # pure continuation row, not the name itself
                    # Strip trailing parenthetical description, whether on the
                    # same line or after an embedded newline (both forms occur).
                    name = re.sub(r"\s*[\(\[].*$", "", text, flags=re.DOTALL).strip()
                    if name:
                        return name
            break
    return None


def locate_columns(header_row):
    cols = {}
    for i, cell in enumerate(header_row):
        if not isinstance(cell, str):
            continue
        for key, pat in COLUMN_NAME_PATTERNS.items():
            if key not in cols and pat.search(cell):
                cols[key] = i
    return cols


def cell(r, cols, key):
    idx = cols.get(key)
    return r[idx] if idx is not None and idx < len(r) else None


def parse_sheet(rows):
    header_idx = next((i for i, r in enumerate(rows) if r and any(
        isinstance(c, str) and COLUMN_NAME_PATTERNS["name"].search(c) for c in r
    )), None)
    if header_idx is None:
        return {"holdings": [], "grand_total": None, "fund_name": None}

    cols = locate_columns(rows[header_idx])
    if "name" not in cols or "weight" not in cols:
        return {"holdings": [], "grand_total": None, "fund_name": None}

    fund_name = find_fund_name(rows)
    if not fund_name:
        return {"holdings": [], "grand_total": None, "fund_name": None}

    holdings = []
    section_label = None
    grand_total = None

    for r in rows[header_idx + 1:]:
        if not r:
            continue
        raw_name = cell(r, cols, "name")
        name = raw_name.strip() if isinstance(raw_name, str) else raw_name
        if not name:
            continue
        if isinstance(name, str) and SKIP_RE.match(name):
            if name.strip().upper() == "GRAND TOTAL":
                grand_total = num(cell(r, cols, "weight"))
                # Everything after GRAND TOTAL is a different table (NAV history,
                # notes, risk-o-meter) whose columns coincidentally hold numeric
                # junk in the same positions — stop here, same fix as PPFAS/HDFC.
                break
            continue

        weight = num(cell(r, cols, "weight"))
        if weight is None:
            # No weight -> section-label row, not a holding. Only overwrite the
            # tracked label when the new text itself matches a known category —
            # nested structural sub-headers ("(a) Listed / awaiting listing...")
            # are reused verbatim across sections and carry no type info of
            # their own (same bug PPFAS/HDFC had to avoid).
            if isinstance(name, str) and any(pat.search(name) for pat, _ in SECTION_TYPE_MAP):
                section_label = name
            continue

        isin = cell(r, cols, "isin")
        industry = cell(r, cols, "industry")
        industry = industry.strip() if isinstance(industry, str) else (industry or "")
        quantity = num(cell(r, cols, "quantity"))
        market_value_lakhs = num(cell(r, cols, "market_value"))
        holdings.append({
            "name": name.strip() if isinstance(name, str) else str(name),
            "isin": (isin.strip() if isinstance(isin, str) else "") or "",
            "industry": industry,
            "quantity": quantity,
            # Lakhs -> Crores (Invesco discloses in Lakhs, same as PPFAS/HDFC).
            "market_value_cr": round(market_value_lakhs / 100, 4) if market_value_lakhs is not None else None,
            # Already a percentage (like HDFC's "% to NAV"), NOT a fraction
            # (unlike PPFAS's "% to Net Assets") — confirmed by inspection.
            "weight": round(weight, 4),
            "type": classify(section_label, name if isinstance(name, str) else ""),
        })

    return {"holdings": holdings, "grand_total": grand_total, "fund_name": fund_name}


def main():
    period = sys.argv[1] if len(sys.argv) > 1 else None
    xlsx_path = Path(sys.argv[2]) if len(sys.argv) > 2 else None
    if not xlsx_path:
        print("usage: parse_invesco_xlsx.py <period YYYY-MM> <path to .xlsx>")
        return 1

    sheets = load_workbook_sheets(xlsx_path)
    funds = {}
    for sheet_name, rows in sheets.items():
        parsed = parse_sheet(rows)
        if not parsed["fund_name"]:
            continue
        total = round(sum(h["weight"] for h in parsed["holdings"]), 2)
        gt = parsed["grand_total"]
        gt_pct = round(gt, 2) if gt is not None else None
        ok = gt_pct is not None and abs(total - gt_pct) < 0.5
        print(f"  {parsed['fund_name']}: {len(parsed['holdings'])} holdings, "
              f"sum={total:.2f}%, GRAND TOTAL={gt_pct}%, {'OK' if ok else 'MISMATCH'}")
        funds[parsed["fund_name"]] = {"holdings": parsed["holdings"], "grand_total_pct": gt_pct}

    if period:
        OUT_DIR.mkdir(parents=True, exist_ok=True)
        out_path = OUT_DIR / f"{period}.json"
        out_path.write_text(json.dumps(funds, indent=2))
        print(f"Written to {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
