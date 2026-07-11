#!/usr/bin/env python3
"""Deterministic parser for Axis Mutual Fund's "Monthly Portfolio Statement"
xls — the SEBI-mandated monthly portfolio disclosure, one file per scheme
(same shape as HDFC's; unlike PPFAS's one-workbook-many-sheets). NOT the
marketing factsheet PDF at
transact.axismf.com/cms/sites/default/files/pdf-factsheets/ — that PDF has no
per-holding ISIN/quantity/weight table, only a curated top-10 list (same
PDF-vs-XLS distinction noted in tools/parse_ppfas_xlsx.py's docstring and this
repo's lib/types.ts header comment). The real disclosure XLS was found via
Axis's own site JSON API (see discover_axis_xlsx.py for the full reverse-
engineering story and its WAF/decoy-response warning) at
www.axismf.com/1/5/.../Monthly_Portfolio_<Scheme>_<date>.xls.

Confirmed by direct inspection of a real downloaded file (Axis Flexi Cap
Fund, June 2026): despite the ".xls" extension the file is a genuine legacy
BIFF8/OLE2 compound document ("Composite Document File V2 Document" magic
bytes) — unlike PPFAS/HDFC where most months are secretly real OOXML zips
renamed to .xls. Every Axis file sampled so far needs xlrd, not openpyxl.
Detect by signature anyway (not extension) in case a future month changes
format, same defensive pattern as the other two parsers.

Layout (one sheet per workbook, named after the AMFI-style scheme code, e.g.
"AXISMLF"):
  row0 col0 = short scheme code (e.g. "AXISMLF"), row0 col1 = full scheme name
  row2ish   = "Monthly Portfolio Statement as on <date>" caption
  header row = col0 (security code, decorative, ignored) | Name of the
    Instrument | ISIN | Industry / Rating | Quantity | Market/Fair Value
    (Rs. in Lakhs) | % to Net Assets | YTM~ | YTC^
  — the SAME 8-visible-column shape as PPFAS's format (locate_columns() finds
  each by header regex, so the extra leading decorative column doesn't need
  special-casing, same as it doesn't for PPFAS).

Row classification: identical "weight presence, not ISIN presence" rule as
PPFAS/HDFC — a row with a "% to Net Assets" value is a real holding (TREPS/
Reverse Repo/Net Current Assets rows have no ISIN but do have a weight); a row
with no weight is a section-label row ("Equity & Equity related", "Debt
Instruments", "Reverse Repo / TREPS", "(a) Listed / awaiting listing on Stock
Exchanges") and updates section-label context for rows beneath it. Same
nested-label fix as PPFAS/HDFC: only overwrite the tracked label when the new
label text itself matches a known category pattern, so a reused structural
sub-header doesn't wipe out the real top-level section's type.

Validation: GRAND TOTAL row's own weight is a fraction of 1 (confirmed:
1.0 for a clean file) — same TRUST_BAND-anchored check as PPFAS/HDFC. Rows
after GRAND TOTAL are derivative-exposure footnotes / disclosure text, not
holdings — stop scanning there, same fix as the other two parsers needed.

Units: Market/Fair Value is in Rs. LAKHS (like HDFC, unlike PPFAS's Crores) —
divide by 100 to get the ₹cr unit lib/types.ts's Holding.market_value and
AnalyseData.aum expect throughout the app. Weight ("% to Net Assets") is a
FRACTION of 1 (like PPFAS, unlike HDFC's already-a-percentage) — multiply by
100.
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

try:
    import xlrd
except ImportError:  # only needed for the legacy-.xls path — see load_workbook_rows
    xlrd = None

ROOT = Path(__file__).parent
CACHE_DIR = ROOT / "cache" / "axis" / "xlsx"
OUT_DIR = ROOT / "out" / "axis_xlsx"

SKIP_RE = re.compile(r"^(sub\s*total|total|grand\s*total)\b", re.IGNORECASE)

SECTION_TYPE_MAP = [
    (re.compile(r"certificate of deposit", re.I), "cd"),
    (re.compile(r"commercial paper", re.I), "cp"),
    (re.compile(r"treasury\s*bill|t-?\s*bills?\b", re.I), "tbill"),
    (re.compile(r"government (bond|security|securities)|g-?sec|state (govern|gover)ment|state development loan", re.I), "gsec"),
    (re.compile(r"treps|reverse repo|repo", re.I), "treps"),
    (re.compile(r"mutual fund", re.I), "fund"),
    (
        re.compile(
            r"corporate (debt|bond)|debenture|\bncd\b|non.convertible|money market|"
            r"\bdebt instrument|\bdebt securit",
            re.I,
        ),
        "corporate_debt",
    ),
    (re.compile(r"cash|net current|net receivable", re.I), "cash"),
    (re.compile(r"reit|invit", re.I), "reit"),
    (re.compile(r"future|option|derivative|hedg", re.I), "derivative"),
    (re.compile(r"preference share", re.I), "preference"),
    (re.compile(r"equity", re.I), "equity"),
]


def classify(section_label: str, name: str = "") -> str:
    for text in (section_label, name):
        for pat, kind in SECTION_TYPE_MAP:
            if pat.search(text or ""):
                return kind
    return "equity"


def num(v):
    if v is None or v == "" or v == "NIL":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


COLUMN_NAME_PATTERNS = {
    "name": re.compile(r"name of the instrument", re.I),
    "isin": re.compile(r"^isin", re.I),
    "industry": re.compile(r"industry|rating", re.I),
    "quantity": re.compile(r"quantity", re.I),
    "market_value": re.compile(r"market.*value", re.I),
    "weight": re.compile(r"%\s*to\s*net", re.I),
}


def load_workbook_rows(path: Path):
    """Every sample inspected so far is genuine legacy BIFF8/OLE2 despite the
    .xls extension (real, not a PPFAS/HDFC-style "secretly OOXML" trick) —
    still detect by file signature rather than trusting the extension, same
    defensive pattern as the other two parsers in case a future month
    switches format."""
    with open(path, "rb") as f:
        sig = f.read(8)
    if sig[:4] == b"PK\x03\x04":
        wb = openpyxl.load_workbook(str(path), data_only=True)
        return {name: [list(row) for row in wb[name].iter_rows(values_only=True)] for name in wb.sheetnames}
    if sig[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        if xlrd is None:
            raise RuntimeError(f"{path} is legacy BIFF8/OLE2 but xlrd isn't installed")
        wb = xlrd.open_workbook(str(path))
        return {name: [wb.sheet_by_name(name).row_values(i) for i in range(wb.sheet_by_name(name).nrows)]
                for name in wb.sheet_names()}
    raise RuntimeError(f"{path}: unrecognized file signature {sig!r}")


def find_fund_name(rows):
    """Row 0 holds a short scheme code in col0 (e.g. "AXISMLF") and the full
    descriptive scheme name in col1 (e.g. "Axis Flexi Cap Fund") — confirmed
    by direct inspection. Fall back to PPFAS's generic "longest cell
    containing 'Fund' in the first few rows" scan if that shape doesn't hold
    (e.g. an ETF/index title without a clean row0/col1, matching HDFC's
    experience that not everything says "Fund")."""
    if rows and len(rows) > 0 and rows[0] and len(rows[0]) > 1:
        cell = rows[0][1]
        if isinstance(cell, str) and cell.strip() and len(cell.strip()) > 3:
            return re.sub(r"\s*[\(\[].*$", "", cell).strip()
    best = None
    for r in rows[:8]:
        for c in r or []:
            if isinstance(c, str) and "fund" in c.lower() and len(c) > 10:
                if best is None or len(c) > len(best):
                    best = c
    if not best:
        return None
    return re.sub(r"\s*[\(\[].*$", "", best).strip()


def locate_columns(header_row):
    cols = {}
    for i, cell in enumerate(header_row):
        if not isinstance(cell, str):
            continue
        for key, pat in COLUMN_NAME_PATTERNS.items():
            if key not in cols and pat.search(cell):
                cols[key] = i
    return cols


def parse_sheet(rows):
    """rows: list[list[cell]] — from either openpyxl or xlrd, via load_workbook_rows()."""
    header_idx = next((i for i, r in enumerate(rows) if r and any(
        isinstance(c, str) and COLUMN_NAME_PATTERNS["name"].search(c) for c in r
    )), None)
    if header_idx is None:
        return {"holdings": [], "grand_total": None, "fund_name": None}

    cols = locate_columns(rows[header_idx])
    if "name" not in cols or "weight" not in cols:
        return {"holdings": [], "grand_total": None, "fund_name": None}

    fund_name = find_fund_name(rows)
    if not fund_name:
        return {"holdings": [], "grand_total": None, "fund_name": None}

    holdings = []
    section_label = None
    grand_total = None

    def cell(r, key):
        idx = cols.get(key)
        return r[idx] if idx is not None and idx < len(r) else None

    for r in rows[header_idx + 1:]:
        if not r:
            continue
        raw_name = cell(r, "name")
        name = raw_name.strip() if isinstance(raw_name, str) else raw_name
        if not name:
            continue
        if isinstance(name, str) and SKIP_RE.match(name):
            if name.strip().upper() == "GRAND TOTAL":
                grand_total = num(cell(r, "weight"))
                # Everything after GRAND TOTAL is derivative-exposure footnote text
                # (confirmed by inspection: rows like "(5) Total outstanding exposure
                # in derivative instruments..." reuse these same column positions with
                # numeric-looking junk) — stop here, same fix PPFAS/HDFC needed.
                break
            continue

        weight_frac = num(cell(r, "weight"))
        if weight_frac is None:
            # Section-label row (no weight at all), not a holding — same rule as
            # PPFAS/HDFC. Only overwrite the tracked label when the new text itself
            # matches a known category, so a reused structural sub-header ("(a)
            # Listed / awaiting listing on Stock Exchanges") doesn't wipe out the
            # real top-level section label the holdings beneath it belong to.
            if isinstance(name, str) and any(pat.search(name) for pat, _ in SECTION_TYPE_MAP):
                section_label = name
            continue

        isin = cell(r, "isin")
        industry = cell(r, "industry")
        industry = industry.strip() if isinstance(industry, str) else (industry or "")
        quantity = num(cell(r, "quantity"))
        market_value_lakhs = num(cell(r, "market_value"))
        holdings.append({
            "name": name.strip() if isinstance(name, str) else str(name),
            "isin": (isin.strip() if isinstance(isin, str) else "") or "",
            "industry": industry,
            "quantity": quantity,
            # Lakhs -> Crores (Axis discloses in Lakhs like HDFC, not Crores like PPFAS).
            "market_value_cr": round(market_value_lakhs / 100, 4) if market_value_lakhs is not None else None,
            # Fraction of 1 (like PPFAS, not already-a-percentage like HDFC) -> *100.
            "weight": round(weight_frac * 100, 4),
            "type": classify(section_label, name if isinstance(name, str) else ""),
        })

    return {"holdings": holdings, "grand_total": grand_total, "fund_name": fund_name}


def main():
    period = sys.argv[1] if len(sys.argv) > 1 else None
    xlsx_path = Path(sys.argv[2]) if len(sys.argv) > 2 else None
    if not xlsx_path:
        print("usage: parse_axis_xlsx.py <period YYYY-MM> <path to .xlsx>")
        return 1

    sheets = load_workbook_rows(xlsx_path)
    funds = {}
    for sheet_name, rows in sheets.items():
        parsed = parse_sheet(rows)
        if not parsed["fund_name"]:
            continue
        total = round(sum(h["weight"] for h in parsed["holdings"]), 2)
        gt = parsed["grand_total"]
        gt_pct = round(gt * 100, 2) if gt is not None else None
        ok = gt_pct is not None and abs(total - gt_pct) < 0.5
        print(f"  {parsed['fund_name']}: {len(parsed['holdings'])} holdings, "
              f"sum={total:.2f}%, GRAND TOTAL={gt_pct}%, {'OK' if ok else 'MISMATCH'}")
        funds[parsed["fund_name"]] = {"holdings": parsed["holdings"], "grand_total_pct": gt_pct}

    if period:
        OUT_DIR.mkdir(parents=True, exist_ok=True)
        out_path = OUT_DIR / f"{period}.json"
        # merge with any funds already parsed for this period (one scheme per file,
        # unlike PPFAS's one-workbook-many-sheets, so periods accumulate incrementally).
        existing = {}
        if out_path.exists():
            existing = json.loads(out_path.read_text())
        existing.update(funds)
        out_path.write_text(json.dumps(existing, indent=2))
        print(f"Written to {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
