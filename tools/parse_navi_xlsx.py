#!/usr/bin/env python3
"""Deterministic parser for Navi Mutual Fund's "Monthly Portfolio Statement"
xlsx — the SEBI-mandated monthly full-portfolio disclosure, one file per
scheme (like HDFC's format, unlike PPFAS's one-workbook-many-sheets). Only
targets the post-rebrand "Navi" era (~2021-09 onward — confirmed by direct
inspection of a June-2025 and April-2026 sample, several scheme types:
Flexi Cap (equity), Liquid Fund (pure debt), Nifty Bank Index Fund
(index/passive)). The pre-rebrand "Essel Mutual Fund" single-workbook era
(2019-01..2021-08) uses a different, incompatible schema and is NOT handled
here — see discover_navi_xlsx.py's docstring; those files are still
downloaded but load_workbook_sheets()/parse_sheet() below will simply find no
"Name of the Instrument" header and skip them (fund_name comes back None),
same graceful-skip behaviour parse_hdfc_xlsx.py relies on for its stray
"Derivative..." disclosure sheets.

Confirmed real OOXML across every sample inspected — openpyxl handles them
directly.

Layout (confirmed by direct inspection, identical across equity/debt/index
scheme types):
  Row 3 (0-indexed), col 1: "NAVI MUTUAL FUND" (fixed banner text)
  Row 4, col 1: the scheme's full name, e.g. "Navi Flexi Cap Fund" — used
    directly, no parenthetical-stripping needed (unlike HDFC's row0/col0
    which has a trailing "(An open ended...)" description; Navi's row4/col1
    is bare). find_fund_name() below still scans defensively (search for
    "Fund" in the first several rows, longest match) rather than hardcoding
    row4/col1, in case older/rarer scheme types shift this — same posture as
    parse_ppfas_xlsx.py's find_fund_name().
  Row 6: header row — "Name of the Instrument" | "ISIN" | "Industry/Rating" |
    "Quantity" | "Market/Fair Value (Rs. in Lacs)" | "% to Net Assets" |
    "YIELD", all starting at col1 (col0 is always blank/decorative, unlike
    HDFC's col0 "|" marker which at least has ink in it).
  Below that: same PPFAS-style dual-purpose column — col1 holds EITHER the
    holding's Name OR (when nothing else in the row is populated) a section
    label ("Equity & Equity related", "(a)Listed / awaiting listing...",
    "Money Market Instruments", "(b)Commercial Paper", etc.) or a Sub Total/
    Total/Grand Total marker. Row-classification rule is the same "weight
    presence, not ISIN presence" test parse_ppfas_xlsx.py and
    parse_hdfc_xlsx.py both use — TREPS/cash summary rows have no ISIN but do
    have a weight, so ISIN presence alone can't be the data-row signal.

Units: "Market/Fair Value (Rs. in Lacs)" — Lacs, not Crores, divide by 100
(same as HDFC; PPFAS's source is already Crores). "% to Net Assets" is
ALREADY a percentage (e.g. 4.27, Grand Total = 100) — same as HDFC, NOT a
fraction like PPFAS's "% to Net Assets"/"% to AUM" column (confirmed:
Grand Total row's own weight cell reads 100, not 1.0) — no *100 scaling.

Navi-specific quirk: a standalone "TREPS/Reverse Repo/Net Current Assets/
Cash/Cash Equivalent" row carries its own weight+market value directly (no
preceding label row of its own) and can appear right after a "(c)Treasury
Bills" sub-section whose label text would otherwise still be the tracked
section_label at that point (nothing resets it at a plain "Total" marker) —
naively classifying by section_label-first (as PPFAS/HDFC do) would tag this
row "tbill" instead of "treps". Harmless for every AGGREGATE the app computes
(both tbill and treps are in DEPLOYABLE_TYPES and both fall in the "Debt"
asset_allocation bucket in build_dataset_navi.py) but wrong as a per-holding
label, so CASH_SUMMARY_RE below checks the row's own name FIRST for this one
unambiguous, recurring, fixed-text case before falling back to PPFAS's
label-first logic for everything else (which still exists specifically to
avoid a company named e.g. "Future Retail Limited" false-matching the
"future" derivative pattern by name).

Validation: same GRAND TOTAL-anchored trust band as PPFAS/HDFC — weight sum
of parsed holdings should land in TRUST_BAND (see parse_all_navi_xlsx.py) and
match the sheet's own Grand Total row within a small tolerance.
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).parent
CACHE_DIR = ROOT / "cache" / "navi" / "xlsx"
OUT_DIR = ROOT / "out" / "navi_xlsx"

SKIP_RE = re.compile(r"^(sub\s*total|total|grand\s*total)\b", re.IGNORECASE)

CASH_SUMMARY_RE = re.compile(r"treps|reverse repo|net current assets|net receivable", re.I)

SECTION_TYPE_MAP = [
    (re.compile(r"certificate of deposit", re.I), "cd"),
    (re.compile(r"commercial paper", re.I), "cp"),
    (re.compile(r"treasury\s*bill|t-?\s*bills?\b", re.I), "tbill"),
    (re.compile(r"gov(?:ern|er)ment (bond|security|securities)|g-?sec|state (govern|gover)ment|state development loan", re.I), "gsec"),
    (re.compile(r"treps|reverse repo|repo", re.I), "treps"),
    (re.compile(r"mutual fund", re.I), "fund"),
    (
        re.compile(
            r"corporate (debt|bond)|debenture|\bncd\b|non.convertible|money market|"
            r"privately placed|securitized debt|\bdebt instrument|\bdebt securit",
            re.I,
        ),
        "corporate_debt",
    ),
    (re.compile(r"cash|net current|net receivable", re.I), "cash"),
    (re.compile(r"reit|invit", re.I), "reit"),
    (re.compile(r"future|option|derivative|hedg", re.I), "derivative"),
    (re.compile(r"preference share", re.I), "preference"),
    (re.compile(r"equity", re.I), "equity"),
]

COLUMN_NAME_PATTERNS = {
    "name": re.compile(r"name of the instrument", re.I),
    "isin": re.compile(r"^isin", re.I),
    "industry": re.compile(r"industry|rating", re.I),
    "quantity": re.compile(r"quantity", re.I),
    "market_value": re.compile(r"market.*value", re.I),
    "weight": re.compile(r"%\s*to\s*net\s*assets", re.I),
}


def classify(section_label: str, name: str = "") -> str:
    # TREPS/cash summary rows are checked by NAME first — see module docstring
    # for why (a stale tracked section_label from a preceding "(c)Treasury
    # Bills" sub-header would otherwise mislabel this fixed, unambiguous line).
    if name and CASH_SUMMARY_RE.search(name):
        return "treps" if re.search(r"treps|reverse repo", name, re.I) else "cash"
    for text in (section_label, name):
        for pat, kind in SECTION_TYPE_MAP:
            if pat.search(text or ""):
                return kind
    return "equity"


def num(v):
    if v is None or v == "" or v == "NIL":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def load_workbook_rows(path: Path):
    with open(path, "rb") as f:
        sig = f.read(4)
    if sig != b"PK\x03\x04":
        raise RuntimeError(f"{path}: not real OOXML (sig={sig!r}) — likely a pre-rebrand Essel-era or .xlsb file, unsupported")
    wb = openpyxl.load_workbook(str(path), data_only=True)
    return {name: [list(row) for row in wb[name].iter_rows(values_only=True)] for name in wb.sheetnames}


def find_fund_name(rows):
    """Row4/col1 is the scheme's bare full name in every sample inspected, but
    scan defensively (longest string containing "Fund" in the first 8 rows)
    rather than hardcoding the position — same posture as
    parse_ppfas_xlsx.py's find_fund_name(), for resilience against layout
    drift across scheme types/years not in the inspected sample set."""
    best = None
    for r in rows[:8]:
        for c in r:
            if isinstance(c, str) and "fund" in c.lower() and len(c) > 6 and c.strip().upper() != "NAVI MUTUAL FUND":
                if best is None or len(c) > len(best):
                    best = c
    if not best:
        return None
    return re.sub(r"\s*\(.*$", "", best).strip()


def locate_columns(header_row):
    cols = {}
    for i, cell in enumerate(header_row):
        if not isinstance(cell, str):
            continue
        for key, pat in COLUMN_NAME_PATTERNS.items():
            if key not in cols and pat.search(cell):
                cols[key] = i
    return cols


def parse_sheet(rows):
    header_idx = next((i for i, r in enumerate(rows) if r and any(
        isinstance(c, str) and COLUMN_NAME_PATTERNS["name"].search(c) for c in r
    )), None)
    if header_idx is None:
        return {"holdings": [], "grand_total": None, "fund_name": None}

    cols = locate_columns(rows[header_idx])
    if "name" not in cols or "weight" not in cols:
        return {"holdings": [], "grand_total": None, "fund_name": None}

    fund_name = find_fund_name(rows)
    if not fund_name:
        return {"holdings": [], "grand_total": None, "fund_name": None}

    holdings = []
    section_label = None
    grand_total = None

    def cell(r, key):
        idx = cols.get(key)
        return r[idx] if idx is not None and idx < len(r) else None

    for r in rows[header_idx + 1:]:
        if not r:
            continue
        raw_name = cell(r, "name")
        name = raw_name.strip() if isinstance(raw_name, str) else raw_name
        if not name:
            continue
        if isinstance(name, str) and SKIP_RE.match(name):
            if name.strip().upper() == "GRAND TOTAL":
                grand_total = num(cell(r, "weight"))
                # Everything after Grand Total is a different table (Plan/Option
                # NAV history, riskometer notes, disclosure footnotes) whose
                # columns coincidentally hold numeric junk in the same positions
                # — stop here, same fix PPFAS/HDFC needed.
                break
            continue

        weight_val = num(cell(r, "weight"))
        if weight_val is None:
            # No weight at all -> section-label row, not a holding. Only overwrite
            # the tracked label when the new text itself matches a known category
            # — nested sub-headers like "(a)Listed / awaiting listing..." are
            # reused verbatim across sections and carry no type info of their own
            # (same nested-label fix PPFAS/HDFC both needed).
            if isinstance(name, str) and any(pat.search(name) for pat, _ in SECTION_TYPE_MAP):
                section_label = name
            continue

        isin = cell(r, "isin")
        industry = cell(r, "industry")
        industry = industry.strip() if isinstance(industry, str) else (industry or "")
        quantity = num(cell(r, "quantity"))
        market_value_lacs = num(cell(r, "market_value"))
        holdings.append({
            "name": name.strip() if isinstance(name, str) else str(name),
            "isin": (isin.strip() if isinstance(isin, str) else "") or "",
            "industry": industry,
            "quantity": quantity,
            "market_value_cr": round(market_value_lacs / 100, 4) if market_value_lacs is not None else None,
            # Already a percentage (Grand Total reads 100, not 1.0) — no *100,
            # unlike PPFAS's fractional "% to Net Assets"/"% to AUM" column.
            "weight": round(weight_val, 4),
            "type": classify(section_label, name if isinstance(name, str) else ""),
        })

    return {"holdings": holdings, "grand_total": grand_total, "fund_name": fund_name}


def main():
    period = sys.argv[1] if len(sys.argv) > 1 else None
    xlsx_path = Path(sys.argv[2]) if len(sys.argv) > 2 else None
    if not xlsx_path:
        print("usage: parse_navi_xlsx.py <period YYYY-MM> <path to .xlsx>")
        return 1

    sheets = load_workbook_rows(xlsx_path)
    funds = {}
    for sheet_name, rows in sheets.items():
        parsed = parse_sheet(rows)
        if not parsed["fund_name"]:
            continue
        total = round(sum(h["weight"] for h in parsed["holdings"]), 2)
        gt = parsed["grand_total"]
        gt_pct = round(gt, 2) if gt is not None else None
        ok = gt_pct is not None and abs(total - gt_pct) < 0.5
        print(f"  {parsed['fund_name']}: {len(parsed['holdings'])} holdings, "
              f"sum={total:.2f}%, GRAND TOTAL={gt_pct}%, {'OK' if ok else 'MISMATCH'}")
        funds[parsed["fund_name"]] = {"holdings": parsed["holdings"], "grand_total_pct": gt_pct}

    if period:
        OUT_DIR.mkdir(parents=True, exist_ok=True)
        out_path = OUT_DIR / f"{period}.json"
        out_path.write_text(json.dumps(funds, indent=2))
        print(f"Written to {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
